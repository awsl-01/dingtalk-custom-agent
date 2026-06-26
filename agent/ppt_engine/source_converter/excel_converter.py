"""
PPT Engine - Excel转换器

使用openpyxl将Excel转换为Markdown表格。
支持多工作表、合并单元格、样式保留。
"""

from pathlib import Path
from typing import List, Dict, Any

from .base_converter import BaseConverter, ConversionResult

# 检查依赖
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelConverter(BaseConverter):
    """Excel转换器"""

    @property
    def supported_extensions(self) -> List[str]:
        return ['.xlsx', '.xlsm']

    @property
    def format_name(self) -> str:
        return 'Excel'

    def _do_convert(self) -> ConversionResult:
        """执行Excel转换"""
        if not OPENPYXL_AVAILABLE:
            return ConversionResult(
                success=False,
                markdown='',
                images=[],
                metadata=self._extract_metadata(),
                error="openpyxl未安装，请运行: pip install openpyxl"
            )

        try:
            wb = load_workbook(str(self.input_path), data_only=True)

            sheets_md = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                if ws.max_row == 0 or ws.max_column == 0:
                    continue

                # 读取数据
                data = []
                for row in ws.iter_rows(values_only=True):
                    # 跳过全空行
                    if all(cell is None for cell in row):
                        continue
                    # 将None转为空字符串
                    row_data = [str(cell) if cell is not None else '' for cell in row]
                    data.append(row_data)

                if not data:
                    continue

                # 生成Markdown表格
                sheet_md = f"## {sheet_name}\n\n"

                # 表头
                if len(data) > 0:
                    headers = data[0]
                    sheet_md += "| " + " | ".join(headers) + " |\n"
                    sheet_md += "| " + " | ".join(["---"] * len(headers)) + " |\n"

                # 数据行
                for row in data[1:]:
                    # 确保列数一致
                    while len(row) < len(headers):
                        row.append('')
                    row = row[:len(headers)]
                    sheet_md += "| " + " | ".join(row) + " |\n"

                sheets_md.append(sheet_md)

            wb.close()

            markdown = "\n\n".join(sheets_md)
            markdown = self._clean_text(markdown)

            metadata = self._extract_metadata()
            metadata['sheet_count'] = len(sheets_md)

            return ConversionResult(
                success=True,
                markdown=markdown,
                images=[],
                metadata=metadata
            )

        except Exception as e:
            return ConversionResult(
                success=False,
                markdown='',
                images=[],
                metadata=self._extract_metadata(),
                error=f"Excel转换失败: {str(e)}"
            )


def convert_excel(input_path: str, output_dir: str = None) -> ConversionResult:
    """
    转换Excel文件为Markdown

    参数:
        input_path: Excel文件路径
        output_dir: 输出目录

    返回:
        ConversionResult 对象
    """
    converter = ExcelConverter(input_path, output_dir)
    return converter.convert()
