"""
巡检核心服务

实现巡检计划管理、打卡、问题上报、工单流转等核心逻辑
数据持久化使用 JSON 文件（与知识库模式一致）
"""
import os
import json
import time
import uuid
import logging
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Tuple
from dataclasses import asdict

from .models import (
    InspectionPlan, InspectionPoint, InspectionRecord,
    InspectionIssue, WorkOrder, InspectionStats,
    CheckItem, CheckResult,
    AreaType, CheckCategory, IssueCategory, IssueStatus,
    PlanStatus, CheckFrequency,
)

logger = logging.getLogger(__name__)


class InspectionService:
    """
    巡检核心服务

    功能：
    1. 计划管理（创建、查询、更新、删除）
    2. 点位管理（创建、查询）
    3. 打卡签到（定位验证、拍照）
    4. 问题上报（图文上报）
    5. 工单流转（派单、整改、复查、关闭）
    6. 统计报表
    """

    def __init__(self, storage_dir: str):
        """
        初始化巡检服务

        参数:
            storage_dir: 数据存储目录
        """
        self._storage_dir = storage_dir
        os.makedirs(storage_dir, exist_ok=True)

        # 数据文件路径
        self._plans_file = os.path.join(storage_dir, "plans.json")
        self._points_file = os.path.join(storage_dir, "points.json")
        self._records_file = os.path.join(storage_dir, "records.json")
        self._issues_file = os.path.join(storage_dir, "issues.json")
        self._orders_file = os.path.join(storage_dir, "orders.json")

        # 内存缓存
        self._plans: Dict[str, InspectionPlan] = {}
        self._points: Dict[str, InspectionPoint] = {}
        self._records: Dict[str, InspectionRecord] = {}
        self._issues: Dict[str, InspectionIssue] = {}
        self._orders: Dict[str, WorkOrder] = {}

        # 加载数据
        self._load_all()

        # 文件修改时间缓存（用于检测外部进程的更新）
        self._file_mtimes: Dict[str, float] = {}
        self._update_mtimes()

    # ==================== 数据持久化 ====================

    def _update_mtimes(self):
        """记录所有数据文件的修改时间"""
        for fpath in [self._plans_file, self._points_file, self._records_file,
                       self._issues_file, self._orders_file]:
            try:
                self._file_mtimes[fpath] = os.path.getmtime(fpath)
            except OSError:
                self._file_mtimes[fpath] = 0.0

    # ==================== 通用更新/批量删除 ====================

    def update_plan(self, plan_id: str, fields: dict) -> bool:
        """更新计划字段"""
        plan = self._plans.get(plan_id)
        if not plan:
            return False
        for key, value in fields.items():
            if hasattr(plan, key) and key not in ("plan_id", "created_at"):
                setattr(plan, key, value)
        plan.updated_at = time.time()
        self._save_plans()
        return True

    def update_point(self, point_id: str, fields: dict) -> bool:
        """更新点位字段"""
        point = self._points.get(point_id)
        if not point:
            return False
        for key, value in fields.items():
            if hasattr(point, key) and key not in ("point_id", "created_at"):
                setattr(point, key, value)
        point.updated_at = time.time()
        self._save_points()
        return True

    def update_record(self, record_id: str, fields: dict) -> bool:
        """更新记录字段"""
        record = self._records.get(record_id)
        if not record:
            return False
        for key, value in fields.items():
            if hasattr(record, key) and key not in ("record_id", "created_at"):
                setattr(record, key, value)
        self._save_records()
        return True

    def update_issue(self, issue_id: str, fields: dict) -> bool:
        """更新问题字段"""
        issue = self._issues.get(issue_id)
        if not issue:
            return False
        for key, value in fields.items():
            if hasattr(issue, key) and key not in ("issue_id", "created_at"):
                setattr(issue, key, value)
        self._save_issues()
        return True

    def update_order(self, order_id: str, fields: dict) -> bool:
        """更新工单字段"""
        order = self._orders.get(order_id)
        if not order:
            return False
        for key, value in fields.items():
            if hasattr(order, key) and key not in ("order_id", "created_at"):
                setattr(order, key, value)
        self._save_orders()
        return True

    def delete_plans_batch(self, plan_ids: list) -> int:
        """批量删除计划"""
        count = 0
        for pid in plan_ids:
            if pid in self._plans:
                del self._plans[pid]
                count += 1
        if count:
            self._save_plans()
        return count

    def delete_points_batch(self, point_ids: list) -> int:
        count = 0
        for pid in point_ids:
            if pid in self._points:
                del self._points[pid]
                count += 1
        if count:
            self._save_points()
        return count

    def delete_records_batch(self, record_ids: list) -> int:
        count = 0
        for rid in record_ids:
            if rid in self._records:
                del self._records[rid]
                count += 1
        if count:
            self._save_records()
        return count

    def delete_issues_batch(self, issue_ids: list) -> int:
        count = 0
        for iid in issue_ids:
            if iid in self._issues:
                del self._issues[iid]
                count += 1
        if count:
            self._save_issues()
        return count

    def delete_orders_batch(self, order_ids: list) -> int:
        count = 0
        for oid in order_ids:
            if oid in self._orders:
                del self._orders[oid]
                count += 1
        if count:
            self._save_orders()
        return count

    def _check_and_reload(self):
        """检查文件是否有外部修改，如有则重新加载"""
        reloaded = False
        for fpath in [self._records_file, self._issues_file, self._orders_file,
                       self._plans_file, self._points_file]:
            try:
                current_mtime = os.path.getmtime(fpath)
                if current_mtime > self._file_mtimes.get(fpath, 0):
                    reloaded = True
                    break
            except OSError:
                pass
        if reloaded:
            logger.info("检测到外部数据修改，重新加载...")
            self._load_all()
            self._update_mtimes()

    def _load_all(self):
        """加载所有数据"""
        self._plans = self._load_json(self._plans_file, InspectionPlan)
        self._points = self._load_json(self._points_file, InspectionPoint)
        self._records = self._load_json(self._records_file, InspectionRecord)
        self._issues = self._load_json(self._issues_file, InspectionIssue)
        self._orders = self._load_json(self._orders_file, WorkOrder)
        logger.info(
            f"加载巡检数据: {len(self._plans)} 计划, {len(self._points)} 点位, "
            f"{len(self._records)} 记录, {len(self._issues)} 问题, {len(self._orders)} 工单"
        )

    def _load_json(self, filepath: str, cls) -> Dict[str, object]:
        """加载 JSON 文件"""
        if not os.path.exists(filepath):
            return {}
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                data = json.load(f)
            return {k: cls(**v) for k, v in data.items()}
        except Exception as e:
            logger.error(f"加载 {filepath} 失败: {e}")
            return {}

    def _save_json(self, filepath: str, data: Dict[str, object]):
        """保存 JSON 文件"""
        try:
            with open(filepath, "w", encoding="utf-8") as f:
                json.dump(
                    {k: asdict(v) for k, v in data.items()},
                    f, ensure_ascii=False, indent=2
                )
        except Exception as e:
            logger.error(f"保存 {filepath} 失败: {e}")

    def _save_plans(self):
        self._save_json(self._plans_file, self._plans)

    def _save_points(self):
        self._save_json(self._points_file, self._points)

    def _save_records(self):
        self._save_json(self._records_file, self._records)

    def _save_issues(self):
        self._save_json(self._issues_file, self._issues)

    def _save_orders(self):
        self._save_json(self._orders_file, self._orders)

    @staticmethod
    def _gen_id(prefix: str) -> str:
        return f"{prefix}_{int(time.time() * 1000)}_{uuid.uuid4().hex[:6]}"

    # ==================== 计划管理 ====================

    def create_plan(
        self,
        plan_name: str,
        area_type: str,
        check_category: str,
        frequency: str,
        assigned_inspectors: list = None,
        assigned_areas: list = None,
        check_items: list = None,
        start_date: str = "",
        end_date: str = "",
        created_by: str = "",
        description: str = "",
    ) -> InspectionPlan:
        """创建巡检计划"""
        plan = InspectionPlan(
            plan_id=self._gen_id("plan"),
            plan_name=plan_name,
            area_type=area_type,
            check_category=check_category,
            frequency=frequency,
            assigned_inspectors=assigned_inspectors or [],
            assigned_areas=assigned_areas or [],
            check_items=check_items or [],
            status=PlanStatus.DRAFT,
            start_date=start_date,
            end_date=end_date,
            created_by=created_by,
            description=description,
        )
        self._plans[plan.plan_id] = plan
        self._save_plans()
        logger.info(f"创建巡检计划: {plan.plan_id} - {plan_name}")
        return plan

    def get_plan(self, plan_id: str) -> Optional[InspectionPlan]:
        return self._plans.get(plan_id)

    def list_plans(
        self,
        status: str = None,
        area_type: str = None,
        inspector_id: str = None,
    ) -> List[InspectionPlan]:
        """查询计划列表"""
        plans = list(self._plans.values())
        if status:
            plans = [p for p in plans if p.status == status]
        if area_type:
            plans = [p for p in plans if p.area_type == area_type]
        if inspector_id:
            plans = [p for p in plans if inspector_id in p.assigned_inspectors]
        plans.sort(key=lambda p: p.created_at, reverse=True)
        return plans

    def update_plan_status(self, plan_id: str, status: str) -> bool:
        plan = self._plans.get(plan_id)
        if not plan:
            return False
        plan.status = status
        plan.updated_at = time.time()
        self._save_plans()
        logger.info(f"更新计划状态: {plan_id} -> {status}")
        return True

    def delete_plan(self, plan_id: str) -> bool:
        if plan_id not in self._plans:
            return False
        del self._plans[plan_id]
        self._save_plans()
        logger.info(f"删除巡检计划: {plan_id}")
        return True

    # ==================== 点位管理 ====================

    def create_point(
        self,
        point_name: str,
        area_type: str,
        location: str,
        check_items: list = None,
        latitude: float = 0.0,
        longitude: float = 0.0,
        requires_photo: bool = True,
        requires_location: bool = False,
        qr_code: str = "",
    ) -> InspectionPoint:
        """创建巡检点位"""
        point = InspectionPoint(
            point_id=self._gen_id("point"),
            point_name=point_name,
            area_type=area_type,
            location=location,
            latitude=latitude,
            longitude=longitude,
            check_items=check_items or [],
            requires_photo=requires_photo,
            requires_location=requires_location,
            qr_code=qr_code,
        )
        self._points[point.point_id] = point
        self._save_points()
        logger.info(f"创建巡检点位: {point.point_id} - {point_name}")
        return point

    def get_point(self, point_id: str) -> Optional[InspectionPoint]:
        return self._points.get(point_id)

    def list_points(
        self,
        area_type: str = None,
        is_active: bool = True,
    ) -> List[InspectionPoint]:
        points = list(self._points.values())
        if area_type:
            points = [p for p in points if p.area_type == area_type]
        if is_active is not None:
            points = [p for p in points if p.is_active == is_active]
        return points

    def delete_point(self, point_id: str) -> bool:
        if point_id not in self._points:
            return False
        del self._points[point_id]
        self._save_points()
        return True

    # ==================== 打卡签到 ====================

    def check_in(
        self,
        plan_id: str,
        point_id: str,
        inspector_id: str,
        inspector_name: str = "",
        latitude: float = 0.0,
        longitude: float = 0.0,
        photo_urls: list = None,
        notes: str = "",
    ) -> Tuple[InspectionRecord, str]:
        """
        巡检打卡签到

        返回:
            (打卡记录, 提示消息)
        """
        plan = self._plans.get(plan_id)
        if not plan:
            return None, "巡检计划不存在"

        if plan.status != PlanStatus.ACTIVE:
            return None, "该计划未激活，无法打卡"

        point = self._points.get(point_id)
        if not point:
            return None, "巡检点位不存在"

        # 定位验证（简单距离校验）
        location_ok = True
        if point.requires_location and point.latitude and point.longitude:
            if latitude and longitude:
                dist = self._calc_distance(
                    latitude, longitude, point.latitude, point.longitude
                )
                # 允许 100 米误差
                location_ok = dist <= 100
            else:
                location_ok = False

        record = InspectionRecord(
            record_id=self._gen_id("record"),
            plan_id=plan_id,
            point_id=point_id,
            inspector_id=inspector_id,
            inspector_name=inspector_name,
            check_in_time=time.time(),
            latitude=latitude,
            longitude=longitude,
            location_verified=location_ok,
            photo_urls=photo_urls or [],
            notes=notes,
        )
        self._records[record.record_id] = record
        self._save_records()

        msg = f"✅ 打卡成功！\n📍 {point.point_name}"
        if not location_ok:
            msg += "\n⚠️ 定位验证未通过，请确认是否在巡检点位附近"
        if point.requires_photo and not photo_urls:
            msg += "\n⚠️ 该点位要求拍照，请补充现场照片"

        logger.info(f"巡检打卡: {inspector_name} -> {point.point_name}")
        return record, msg

    def check_out(
        self,
        record_id: str,
        check_results: list = None,
    ) -> Tuple[bool, str]:
        """
        巡检签退（提交检查结果）

        参数:
            record_id: 记录ID
            check_results: 检查结果列表 [{"item_id": "", "status": "pass/warn/fail", "description": ""}]

        返回:
            (是否成功, 提示消息)
        """
        record = self._records.get(record_id)
        if not record:
            return False, "打卡记录不存在"

        record.check_out_time = time.time()
        record.check_results = check_results or []

        # 判断是否有问题
        has_issues = any(
            r.get("status") == "fail" for r in record.check_results
        )
        record.overall_status = "has_issues" if has_issues else "normal"

        self._save_records()

        if has_issues:
            return True, "⚠️ 检查发现异常项，请在问题上报中补充详情"
        return True, "✅ 巡检完成，所有检查项正常"

    # ==================== 问题上报 ====================

    def report_issue(
        self,
        record_id: str,
        category: str,
        title: str,
        description: str,
        reported_by: str,
        reported_by_name: str = "",
        photo_urls: list = None,
        severity: str = "medium",
        point_name: str = "",
    ) -> InspectionIssue:
        """上报巡检问题"""
        record = self._records.get(record_id)
        plan_id = record.plan_id if record else ""

        issue = InspectionIssue(
            issue_id=self._gen_id("issue"),
            record_id=record_id,
            plan_id=plan_id,
            point_id=record.point_id if record else "",
            point_name=point_name,
            category=category,
            title=title,
            description=description,
            photo_urls=photo_urls or [],
            reported_by=reported_by,
            reported_by_name=reported_by_name,
            status=IssueStatus.PENDING,
            severity=severity,
        )
        self._issues[issue.issue_id] = issue
        self._save_issues()

        # 自动创建工单
        self._create_workorder_for_issue(issue)

        logger.info(f"上报问题: {issue.issue_id} - {title}")
        return issue

    def _create_workorder_for_issue(self, issue: InspectionIssue):
        """为问题自动创建工单"""
        order = WorkOrder(
            order_id=self._gen_id("order"),
            issue_id=issue.issue_id,
            status=IssueStatus.PENDING,
        )
        self._orders[order.order_id] = order
        self._save_orders()

    # ==================== 工单流转 ====================

    def assign_order(
        self,
        issue_id: str,
        assigned_to: str,
        assigned_to_name: str = "",
        assigned_by: str = "",
        deadline_hours: int = 24,
    ) -> Tuple[bool, str]:
        """派单：将问题分配给整改负责人"""
        issue = self._issues.get(issue_id)
        if not issue:
            return False, "问题不存在"

        if issue.status not in (IssueStatus.PENDING, IssueStatus.REJECTED):
            return False, f"当前状态({issue.status})无法派单"

        issue.assigned_to = assigned_to
        issue.assigned_to_name = assigned_to_name
        issue.assigned_at = time.time()
        issue.deadline = time.time() + deadline_hours * 3600
        issue.status = IssueStatus.ASSIGNED

        # 更新工单
        for order in self._orders.values():
            if order.issue_id == issue_id:
                order.assigned_to = assigned_to
                order.assigned_to_name = assigned_to_name
                order.assigned_by = assigned_by
                order.deadline = issue.deadline
                order.status = IssueStatus.ASSIGNED
                order.updated_at = time.time()
                order.operations.append({
                    "action": "assign",
                    "operator": assigned_by,
                    "target": assigned_to,
                    "timestamp": time.time(),
                })
                break

        self._save_issues()
        self._save_orders()
        logger.info(f"派单: {issue_id} -> {assigned_to_name}")
        return True, f"✅ 已派单给 {assigned_to_name}，整改期限 {deadline_hours} 小时"

    def start_rectification(
        self,
        issue_id: str,
        operator: str = "",
    ) -> Tuple[bool, str]:
        """开始整改"""
        issue = self._issues.get(issue_id)
        if not issue:
            return False, "问题不存在"

        if issue.status != IssueStatus.ASSIGNED:
            return False, f"当前状态({issue.status})无法开始整改"

        issue.status = IssueStatus.IN_PROGRESS
        self._update_order_status(issue_id, IssueStatus.IN_PROGRESS, operator)
        self._save_issues()
        self._save_orders()
        return True, "✅ 已开始整改"

    def submit_rectification(
        self,
        issue_id: str,
        rectification_photos: list = None,
        rectification_notes: str = "",
        operator: str = "",
    ) -> Tuple[bool, str]:
        """提交整改结果，申请复查"""
        issue = self._issues.get(issue_id)
        if not issue:
            return False, "问题不存在"

        if issue.status not in (IssueStatus.IN_PROGRESS, IssueStatus.ASSIGNED):
            return False, f"当前状态({issue.status})无法提交整改"

        issue.rectification_photos = rectification_photos or []
        issue.rectification_notes = rectification_notes
        issue.rectified_at = time.time()
        issue.status = IssueStatus.PENDING_REVIEW
        self._update_order_status(issue_id, IssueStatus.PENDING_REVIEW, operator)
        self._save_issues()
        self._save_orders()
        return True, "✅ 整改已提交，等待复查验收"

    def review_issue(
        self,
        issue_id: str,
        review_result: str,  # pass/reject
        reviewer_id: str,
        review_notes: str = "",
    ) -> Tuple[bool, str]:
        """复查验收"""
        issue = self._issues.get(issue_id)
        if not issue:
            return False, "问题不存在"

        if issue.status != IssueStatus.PENDING_REVIEW:
            return False, f"当前状态({issue.status})无法复查"

        issue.reviewer_id = reviewer_id
        issue.review_result = review_result
        issue.review_notes = review_notes
        issue.reviewed_at = time.time()

        if review_result == "pass":
            issue.status = IssueStatus.RESOLVED
            self._update_order_status(issue_id, IssueStatus.RESOLVED, reviewer_id)
            msg = "✅ 复查通过，问题已关闭"
        else:
            issue.status = IssueStatus.REJECTED
            self._update_order_status(issue_id, IssueStatus.REJECTED, reviewer_id)
            msg = "❌ 复查未通过，已退回整改"

        self._save_issues()
        self._save_orders()
        return True, msg

    def _update_order_status(self, issue_id: str, status: str, operator: str):
        """更新工单状态"""
        for order in self._orders.values():
            if order.issue_id == issue_id:
                order.status = status
                order.updated_at = time.time()
                order.operations.append({
                    "action": status,
                    "operator": operator,
                    "timestamp": time.time(),
                })
                break

    # ==================== 查询 ====================

    def get_issue(self, issue_id: str) -> Optional[InspectionIssue]:
        return self._issues.get(issue_id)

    def list_issues(
        self,
        status: str = None,
        category: str = None,
        severity: str = None,
        assigned_to: str = None,
        plan_id: str = None,
    ) -> List[InspectionIssue]:
        issues = list(self._issues.values())
        if status:
            issues = [i for i in issues if i.status == status]
        if category:
            issues = [i for i in issues if i.category == category]
        if severity:
            issues = [i for i in issues if i.severity == severity]
        if assigned_to:
            issues = [i for i in issues if i.assigned_to == assigned_to]
        if plan_id:
            issues = [i for i in issues if i.plan_id == plan_id]
        issues.sort(key=lambda i: i.created_at, reverse=True)
        return issues

    def list_records(
        self,
        plan_id: str = None,
        point_id: str = None,
        inspector_id: str = None,
        date_str: str = None,
    ) -> List[InspectionRecord]:
        records = list(self._records.values())
        if plan_id:
            records = [r for r in records if r.plan_id == plan_id]
        if point_id:
            records = [r for r in records if r.point_id == point_id]
        if inspector_id:
            records = [r for r in records if r.inspector_id == inspector_id]
        if date_str:
            # 验证日期格式
            import re
            date_pattern = re.compile(r'^\d{4}-\d{2}-\d{2}$')
            if not date_pattern.match(date_str):
                logger.warning(f"无效的日期格式: {date_str}")
                return []
            try:
                dt = datetime.fromisoformat(date_str)
                start = dt.replace(hour=0, minute=0, second=0).timestamp()
                end = dt.replace(hour=23, minute=59, second=59).timestamp()
                records = [r for r in records if start <= r.check_in_time <= end]
            except ValueError as e:
                logger.error(f"日期解析失败: {date_str}, 错误: {e}")
                return []
        records.sort(key=lambda r: r.check_in_time, reverse=True)
        return records

    def get_orders(
        self,
        status: str = None,
        assigned_to: str = None,
    ) -> List[WorkOrder]:
        orders = list(self._orders.values())
        if status:
            orders = [o for o in orders if o.status == status]
        if assigned_to:
            orders = [o for o in orders if o.assigned_to == assigned_to]
        orders.sort(key=lambda o: o.created_at, reverse=True)
        return orders

    # ==================== 统计 ====================

    def get_stats(self) -> InspectionStats:
        """获取巡检统计"""
        now = time.time()
        today_start = datetime.now().replace(hour=0, minute=0, second=0).timestamp()
        week_start = (datetime.now() - timedelta(days=7)).timestamp()

        records_today = sum(
            1 for r in self._records.values()
            if r.check_in_time >= today_start
        )
        records_week = sum(
            1 for r in self._records.values()
            if r.check_in_time >= week_start
        )

        issues = list(self._issues.values())
        by_area = {}
        by_category = {}
        by_severity = {}
        for issue in issues:
            by_area[issue.point_name] = by_area.get(issue.point_name, 0) + 1
            by_category[issue.category] = by_category.get(issue.category, 0) + 1
            by_severity[issue.severity] = by_severity.get(issue.severity, 0) + 1

        return InspectionStats(
            total_plans=len(self._plans),
            active_plans=sum(1 for p in self._plans.values() if p.status == PlanStatus.ACTIVE),
            total_points=len(self._points),
            total_records_today=records_today,
            total_records_week=records_week,
            total_issues=len(issues),
            pending_issues=sum(1 for i in issues if i.status in (IssueStatus.PENDING, IssueStatus.ASSIGNED, IssueStatus.IN_PROGRESS)),
            resolved_issues=sum(1 for i in issues if i.status in (IssueStatus.RESOLVED, IssueStatus.CLOSED)),
            total_orders=len(self._orders),
            pending_orders=sum(1 for o in self._orders.values() if o.status not in (IssueStatus.RESOLVED, IssueStatus.CLOSED)),
            by_area=by_area,
            by_category=by_category,
            by_severity=by_severity,
        )

    # ==================== 工具 ====================

    # ==================== 数据导入 ====================

    def import_points_from_excel(self, file_path: str) -> Tuple[int, int, str]:
        """
        从 Excel 文件导入巡检点位

        参数:
            file_path: Excel 文件路径

        返回:
            (成功数, 跳过数, 提示消息)
        """
        try:
            import openpyxl
        except ImportError:
            return 0, 0, "需要安装 openpyxl: pip install openpyxl"

        if not os.path.exists(file_path):
            return 0, 0, f"文件不存在: {file_path}"

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active

            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                return 0, 0, "文件为空或只有表头"

            # 解析表头，获取列索引
            header = [str(h).strip() if h else "" for h in rows[0]]
            col_map = self._parse_column_map(header)

            if "point_name" not in col_map:
                return 0, 0, "未找到[点位名称]列，请检查表头"

            imported = 0
            skipped = 0
            errors = []

            for i, row in enumerate(rows[1:], start=2):
                try:
                    point_name = self._get_cell_value(row, col_map.get("point_name", -1))
                    if not point_name:
                        skipped += 1
                        continue

                    area_type = self._get_cell_value(row, col_map.get("area_type", -1)) or "public"
                    location = self._get_cell_value(row, col_map.get("location", -1)) or ""

                    # 标准化区域类型
                    area_type = self._normalize_area_type(area_type)

                    # 解析经纬度
                    lat = self._parse_float(self._get_cell_value(row, col_map.get("latitude", -1)))
                    lon = self._parse_float(self._get_cell_value(row, col_map.get("longitude", -1)))

                    # 解析是否要求拍照/定位
                    requires_photo = self._parse_bool(self._get_cell_value(row, col_map.get("requires_photo", -1)))
                    requires_location = self._parse_bool(self._get_cell_value(row, col_map.get("requires_location", -1)))

                    # 解析检查项
                    check_items_str = self._get_cell_value(row, col_map.get("check_items", -1)) or ""
                    check_items = self._parse_check_items(check_items_str)

                    # 解析检查人和检查时间
                    inspector = self._get_cell_value(row, col_map.get("inspector", -1)) or ""
                    check_time = self._get_cell_value(row, col_map.get("check_time", -1)) or ""

                    # 创建点位
                    point = self.create_point(
                        point_name=point_name,
                        area_type=area_type,
                        location=location,
                        check_items=check_items,
                        latitude=lat,
                        longitude=lon,
                        requires_photo=requires_photo,
                        requires_location=requires_location,
                    )

                    # 存储检查人和检查时间到 metadata
                    if inspector or check_time:
                        point.metadata["inspector"] = inspector
                        point.metadata["check_time"] = check_time
                        self._save_points()
                    imported += 1

                except Exception as e:
                    errors.append(f"第{i}行: {str(e)}")
                    skipped += 1

            wb.close()

            msg = f"✅ 导入完成：成功 {imported} 个，跳过 {skipped} 个"
            if errors:
                msg += f"\n⚠️ 错误：{'; '.join(errors[:3])}"

            logger.info(f"导入巡检点位: 成功 {imported}, 跳过 {skipped}")
            return imported, skipped, msg

        except Exception as e:
            logger.error(f"导入 Excel 失败: {e}")
            return 0, 0, f"导入失败: {str(e)}"

    def import_issues_from_excel(self, file_path: str) -> Tuple[int, int, str]:
        """
        从 Excel 文件导入巡检问题

        参数:
            file_path: Excel 文件路径

        返回:
            (成功数, 跳过数, 提示消息)
        """
        try:
            import openpyxl
        except ImportError:
            return 0, 0, "需要安装 openpyxl: pip install openpyxl"

        if not os.path.exists(file_path):
            return 0, 0, f"文件不存在: {file_path}"

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active

            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                return 0, 0, "文件为空或只有表头"

            header = [str(h).strip() if h else "" for h in rows[0]]
            col_map = self._parse_column_map(header)

            if "title" not in col_map:
                return 0, 0, "未找到[问题标题]列，请检查表头"

            imported = 0
            skipped = 0
            errors = []

            for i, row in enumerate(rows[1:], start=2):
                try:
                    title = self._get_cell_value(row, col_map.get("title", -1))
                    if not title:
                        skipped += 1
                        continue

                    category = self._get_cell_value(row, col_map.get("category", -1)) or "other"
                    severity = self._get_cell_value(row, col_map.get("severity", -1)) or "medium"
                    description = self._get_cell_value(row, col_map.get("description", -1)) or ""
                    point_name = self._get_cell_value(row, col_map.get("point_name", -1)) or ""
                    area_type = self._get_cell_value(row, col_map.get("area_type", -1)) or ""
                    reported_by_name = self._get_cell_value(row, col_map.get("inspector", -1)) or "Excel导入"
                    reported_time = self._get_cell_value(row, col_map.get("check_time", -1)) or ""

                    # 标准化
                    category = self._normalize_issue_category(category)
                    severity = self._normalize_severity(severity)

                    issue = self.report_issue(
                        record_id="",
                        category=category,
                        title=title,
                        description=description,
                        reported_by="import",
                        reported_by_name=reported_by_name,
                        severity=severity,
                        point_name=point_name,
                    )

                    # 存储上报时间到 metadata
                    if reported_time:
                        issue.metadata = {"reported_time": reported_time}
                        self._save_issues()
                    imported += 1

                except Exception as e:
                    errors.append(f"第{i}行: {str(e)}")
                    skipped += 1

            wb.close()

            msg = f"✅ 导入完成：成功 {imported} 个，跳过 {skipped} 个"
            if errors:
                msg += f"\n⚠️ 错误：{'; '.join(errors[:3])}"

            logger.info(f"导入巡检问题: 成功 {imported}, 跳过 {skipped}")
            return imported, skipped, msg

        except Exception as e:
            logger.error(f"导入 Excel 失败: {e}")
            return 0, 0, f"导入失败: {str(e)}"

    def _parse_column_map(self, header: list) -> dict:
        """解析表头，建立列名到索引的映射"""
        col_map = {}
        # 中文表头到字段名的映射
        name_map = {
            "点位名称": "point_name", "名称": "point_name",
            "区域类型": "area_type", "区域": "area_type",
            "位置描述": "location", "位置": "location",
            "纬度": "latitude", "经度": "longitude",
            "是否要求拍照": "requires_photo", "拍照": "requires_photo",
            "是否要求定位": "requires_location", "定位": "requires_location",
            "检查项": "check_items",
            "检查人": "inspector", "负责老师": "inspector", "检查老师": "inspector",
            "检查时间": "check_time", "巡检时间": "check_time",
            "问题标题": "title", "标题": "title",
            "问题分类": "category", "分类": "category",
            "严重程度": "severity",
            "问题描述": "description", "描述": "description",
        }
        for idx, h in enumerate(header):
            h_lower = h.lower().strip()
            # 先精确匹配英文
            if h_lower in ("point_name", "area_type", "location", "latitude", "longitude",
                           "requires_photo", "requires_location", "check_items",
                           "title", "category", "severity", "description"):
                col_map[h_lower] = idx
            # 再匹配中文
            elif h in name_map:
                col_map[name_map[h]] = idx
        return col_map

    @staticmethod
    def _get_cell_value(row: tuple, idx: int) -> str:
        if idx < 0 or idx >= len(row):
            return ""
        val = row[idx]
        if val is None:
            return ""
        return str(val).strip()

    @staticmethod
    def _parse_float(val: str) -> float:
        try:
            return float(val) if val else 0.0
        except (ValueError, TypeError):
            return 0.0

    @staticmethod
    def _parse_bool(val: str) -> bool:
        return val in ("是", "yes", "true", "1", "True")

    @staticmethod
    def _normalize_area_type(val: str) -> str:
        mapping = {
            "teaching": "teaching", "教学区": "teaching", "教学": "teaching",
            "dormitory": "dormitory", "宿舍区": "dormitory", "宿舍": "dormitory",
            "canteen": "canteen", "食堂": "canteen",
            "playground": "playground", "操场": "playground",
            "fire": "fire", "消防": "fire", "消防设施": "fire",
            "public": "public", "公共区域": "public", "公共": "public",
        }
        return mapping.get(val.lower().strip(), "public")

    @staticmethod
    def _normalize_issue_category(val: str) -> str:
        mapping = {
            "safety_hazard": "safety_hazard", "安全隐患": "safety_hazard", "安全": "safety_hazard",
            "hygiene_issue": "hygiene_issue", "卫生问题": "hygiene_issue", "卫生": "hygiene_issue",
            "facility_damage": "facility_damage", "设施损坏": "facility_damage", "设施": "facility_damage",
            "discipline_violation": "discipline_violation", "纪律违规": "discipline_violation",
            "fire_safety": "fire_safety", "消防安全": "fire_safety", "消防": "fire_safety",
            "other": "other", "其他": "other",
        }
        return mapping.get(val.lower().strip(), "other")

    @staticmethod
    def _normalize_severity(val: str) -> str:
        mapping = {
            "low": "low", "低": "low",
            "medium": "medium", "中": "medium",
            "high": "high", "高": "high",
            "critical": "critical", "严重": "critical",
        }
        return mapping.get(val.lower().strip(), "medium")

    @staticmethod
    def _parse_check_items(items_str: str) -> list:
        """解析检查项字符串"""
        if not items_str:
            return []
        items = [item.strip() for item in items_str.split(",") if item.strip()]
        return [
            {"item_id": f"item_{i}", "item_name": name, "category": "general"}
            for i, name in enumerate(items)
        ]

    @staticmethod
    def _calc_distance(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
        """计算两点之间的距离（米），使用 Haversine 公式"""
        import math
        R = 6371000  # 地球半径（米）
        phi1 = math.radians(lat1)
        phi2 = math.radians(lat2)
        dphi = math.radians(lat2 - lat1)
        dlambda = math.radians(lon2 - lon1)
        a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
        return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


# ==================== 全局实例 ====================

_service: Optional[InspectionService] = None


def get_inspection_service(storage_dir: str = None) -> InspectionService:
    """获取全局巡检服务实例"""
    global _service
    if _service is None:
        if storage_dir is None:
            storage_dir = os.path.join("knowledge", "inspection")
        _service = InspectionService(storage_dir)
    else:
        # 检查外部进程（钉钉机器人）是否修改了数据文件
        _service._check_and_reload()
    return _service
