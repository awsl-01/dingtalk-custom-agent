"""
巡检技能 - 通过钉钉指令触发巡检打卡、问题上报、工单处理等

支持的指令：
- 巡检打卡 [点位名称]         → 打卡签到
- 巡检签退 [记录ID]           → 签退并提交检查结果
- 上报问题 [描述]             → 上报巡检发现的问题
- 巡检计划                    → 查看巡检计划列表
- 巡检统计                    → 查看巡检统计
- 巡检记录 [日期]             → 查看巡检记录
- 补充照片 [记录ID]           → 为打卡记录补充照片
- 工单列表                    → 查看待处理工单
- 整改完成 [问题ID]           → 提交整改结果
- 复查 [问题ID] 通过/不通过    → 复查验收
- 巡检点位模板               → 下载点位导入模板
- 巡检问题模板               → 下载问题上报模板
"""
import os
import csv
import time
import hashlib
import logging
from typing import List, Optional
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from agent.skills.registry import BaseSkill, skill_registry

logger = logging.getLogger(__name__)

# 照片存储目录
PHOTOS_DIR = os.path.join("knowledge", "inspection", "photos")


class InspectionSkill(BaseSkill):
    """巡检技能"""

    @property
    def name(self) -> str:
        return "巡检管理"

    @property
    def description(self) -> str:
        return "学校区域巡检：打卡签到、问题上报、工单处理、统计查看"

    @property
    def keywords(self) -> List[str]:
        return [
            "巡检", "打卡", "签到", "签退", "点位",
            "上报问题", "巡检计划", "巡检统计", "巡检记录",
            "补充照片", "用户记录", "查看记录", "点位记录", "工单", "整改", "复查",
            "点位模板", "问题模板",
            "确认接收", "接收工单", "工单完成", "维修完成",
        ]

    @property
    def priority(self) -> int:
        return 60

    def can_handle(self, text: str) -> float:
        """判断是否能处理此消息"""
        text_lower = text.lower().strip()

        # 高置信度匹配
        high_keywords = ["巡检打卡", "巡检签退", "上报问题", "巡检计划", "巡检统计", "工单列表", "整改完成", "点位模板", "问题模板", "补充照片", "用户记录", "查看记录", "点位记录", "确认接收", "接收工单", "工单完成", "维修完成"]
        for kw in high_keywords:
            if kw in text_lower:
                return 0.95

        # 中等置信度匹配
        mid_keywords = ["打卡", "签到", "签退", "巡检", "工单", "整改", "复查", "照片"]
        for kw in mid_keywords:
            if kw in text_lower:
                return 0.8

        return 0

    def extract_info(self, text: str) -> dict:
        """从消息中提取信息 - 支持LLM解析和规则回退"""
        text_lower = text.lower().strip()
        info = {"action": "", "params": ""}

        import re

        # 优先检查明确的指令关键词（高优先级）
        # 巡检记录必须在照片特殊模式之前检查
        if "巡检记录" in text_lower:
            info["action"] = "list_records"

            # 尝试使用LLM解析自然语言查询
            llm_result = self._try_llm_parse(text)
            if llm_result and llm_result.get("confidence", 0) >= 0.8:
                # 如果LLM解析出了日期，使用LLM的结果
                if llm_result.get("params"):
                    info["params"] = llm_result["params"]
                logger.info(f"使用LLM解析巡检记录: {info}")
                return info

            # 回退到规则解析
            # 支持多种格式：
            # - "巡检记录 2026-06-22"
            # - "昨天的巡检记录"
            # - "巡检记录 操场看台"
            for prefix in ["巡检记录"]:
                if prefix in text_lower:
                    # 提取"巡检记录"前后的所有内容
                    parts = text_lower.split(prefix, 1)
                    before_prefix = parts[0].strip() if parts[0] else ""
                    after_prefix = parts[1].strip() if len(parts) > 1 else ""

                    # 合并前后内容（去掉"的"字）
                    all_params = []
                    if before_prefix:
                        # 移除末尾的"的"字
                        if before_prefix.endswith("的"):
                            before_prefix = before_prefix[:-1]
                        if before_prefix:
                            all_params.append(before_prefix)
                    if after_prefix:
                        all_params.append(after_prefix)

                    info["params"] = " ".join(all_params) if all_params else ""
                    break

            return info

        # 特殊模式：点位名称 + 巡检照片（如"操场看台 巡检照片"或"操场看台 巡检照片 2026-06-16"）
        # 这种格式应该等同于 "点位记录 点位名称"
        photo_pattern = re.match(r'^(.+?)\s*巡检照片(?:\s+(\d{4}-\d{2}-\d{2}))?$', text_lower)
        if photo_pattern:
            point_name = photo_pattern.group(1).strip()
            date_str = photo_pattern.group(2) or ""
            if point_name:
                params = point_name
                if date_str:
                    params = f"{point_name} {date_str}"
                info["action"] = "point_records"
                info["params"] = params
                return info

        # 特殊模式：点位名称 + 照片（如"操场看台 照片"或"操场看台 照片 2026-06-16"）
        # 但排除包含"巡检记录"、"点位记录"、"用户记录"等指令的消息
        has_record_keyword = any(kw in text_lower for kw in ["巡检记录", "点位记录", "用户记录", "查看记录"])
        if not has_record_keyword:
            photo_pattern2 = re.match(r'^(.+?)\s*照片(?:\s+(\d{4}-\d{2}-\d{2}))?$', text_lower)
            if photo_pattern2:
                point_name = photo_pattern2.group(1).strip()
                date_str = photo_pattern2.group(2) or ""
                if point_name:
                    params = point_name
                    if date_str:
                        params = f"{point_name} {date_str}"
                    info["action"] = "point_records"
                    info["params"] = params
                    return info

        # 尝试使用LLM解析自然语言查询
        llm_result = self._try_llm_parse(text)
        if llm_result and llm_result.get("confidence", 0) >= 0.8:
            return llm_result

        # 回退到规则解析
        return self._rule_based_parse(text_lower, info)

    def _try_llm_parse(self, text: str) -> dict:
        """尝试使用LLM解析用户意图"""
        try:
            from agent.llm_intent_parser import get_llm_intent_parser
            parser = get_llm_intent_parser()
            llm_result = parser.parse_inspection_query(text)

            if llm_result and llm_result.get("intent") == "query_inspection":
                # 转换为巡检技能的格式
                info = {"action": "", "params": ""}

                # 兼容新旧字段名
                query_type = llm_result.get("type") or llm_result.get("query_type", "records")
                want_photos = llm_result.get("photos") or llm_result.get("want_photos", False)

                if query_type == "photos" or want_photos:
                    info["action"] = "point_records"
                elif query_type == "stats":
                    info["action"] = "stats"
                else:
                    info["action"] = "list_records"

                # 构建参数 - 兼容新旧字段名
                params_parts = []
                point_name = llm_result.get("point") or llm_result.get("point_name")
                if point_name:
                    params_parts.append(point_name)
                if llm_result.get("date"):
                    params_parts.append(llm_result["date"])
                else:
                    date_rel = llm_result.get("date_rel") or llm_result.get("date_relative")
                    if date_rel:
                        params_parts.append(date_rel)

                info["params"] = " ".join(params_parts) if params_parts else ""
                info["llm_parsed"] = True
                info["confidence"] = llm_result.get("confidence", 0.8)

                logger.info(f"LLM解析巡检意图成功: {info}")
                return info

        except Exception as e:
            logger.warning(f"LLM解析失败，回退到规则解析: {e}")

        return None

    def _rule_based_parse(self, text_lower: str, info: dict) -> dict:
        """规则解析（回退方案）"""
        import re

        if "巡检打卡" in text_lower or "打卡" in text_lower or "签到" in text_lower:
            info["action"] = "check_in"
            # 提取点位名称（"巡检打卡 教学楼A"）
            for prefix in ["巡检打卡", "打卡", "签到"]:
                if prefix in text_lower:
                    params = text_lower.split(prefix, 1)
                    if len(params) > 1:
                        info["params"] = params[1].strip()
                    break

        elif "巡检签退" in text_lower or "签退" in text_lower:
            info["action"] = "check_out"
            for prefix in ["巡检签退", "签退"]:
                if prefix in text_lower:
                    params = text_lower.split(prefix, 1)
                    if len(params) > 1:
                        info["params"] = params[1].strip()
                    break

        elif "上报问题" in text_lower:
            info["action"] = "report_issue"
            params = text_lower.split("上报问题", 1)
            if len(params) > 1:
                info["params"] = params[1].strip()

        elif "巡检计划" in text_lower:
            info["action"] = "list_plans"

        elif "巡检统计" in text_lower:
            info["action"] = "stats"

        elif "巡检记录" in text_lower:
            info["action"] = "list_records"
            # 提取参数（支持多种格式）
            for prefix in ["巡检记录"]:
                if prefix in text_lower:
                    params = text_lower.split(prefix, 1)
                    if len(params) > 1:
                        info["params"] = params[1].strip()
                    break

        elif "用户记录" in text_lower or "查看记录" in text_lower:
            info["action"] = "user_records"
            # 提取参数（支持"用户记录 启拓"或"用户记录 启拓 操场看台"）
            for prefix in ["用户记录", "查看记录"]:
                if prefix in text_lower:
                    params = text_lower.split(prefix, 1)
                    if len(params) > 1:
                        info["params"] = params[1].strip()
                    break

        elif "点位记录" in text_lower:
            info["action"] = "point_records"
            # 提取点位名称（如"点位记录 操场看台"）
            for prefix in ["点位记录"]:
                if prefix in text_lower:
                    params = text_lower.split(prefix, 1)
                    if len(params) > 1:
                        info["params"] = params[1].strip()
                    break

        elif "补充照片" in text_lower:
            info["action"] = "add_photo"
            # 提取记录ID（如"补充照片 record_xxx"）
            for prefix in ["补充照片"]:
                if prefix in text_lower:
                    params = text_lower.split(prefix, 1)
                    if len(params) > 1:
                        info["params"] = params[1].strip()
                    break

        elif "工单列表" in text_lower or "我的工单" in text_lower:
            info["action"] = "my_orders"

        elif "整改完成" in text_lower:
            info["action"] = "submit_rectification"
            params = text_lower.split("整改完成", 1)
            if len(params) > 1:
                info["params"] = params[1].strip()

        elif "复查" in text_lower:
            info["action"] = "review"
            # 提取问题ID和结果
            parts = text_lower.split("复查", 1)
            if len(parts) > 1:
                rest = parts[1].strip()
                if "通过" in rest:
                    info["params"] = {"result": "pass"}
                elif "不通过" in rest or "退回" in rest:
                    info["params"] = {"result": "reject"}
                    # 提取问题ID
                    for token in rest.split():
                        if token.startswith("issue_") or token.startswith("问题"):
                            info["params"]["issue_id"] = token.replace("问题", "")

        elif "点位模板" in text_lower or "点位导入" in text_lower:
            info["action"] = "point_template"

        elif "问题模板" in text_lower or "问题上报模板" in text_lower:
            info["action"] = "issue_template"

        # 新增：负责人确认接收工单
        elif "确认接收" in text_lower or "接收工单" in text_lower:
            info["action"] = "accept_order"
            # 提取工单号（如"确认接收 issue_xxx"或"接收工单 issue_xxx"）
            for prefix in ["确认接收", "接收工单"]:
                if prefix in text_lower:
                    params = text_lower.split(prefix, 1)
                    if len(params) > 1:
                        info["params"] = params[1].strip()
                    break

        # 新增：工单完成确认
        elif "工单完成" in text_lower or "维修完成" in text_lower:
            info["action"] = "complete_order"
            # 提取工单号
            for prefix in ["工单完成", "维修完成"]:
                if prefix in text_lower:
                    params = text_lower.split(prefix, 1)
                    if len(params) > 1:
                        info["params"] = params[1].strip()
                    break

        return info

    async def execute(self, text: str, context: dict) -> str:
        """执行巡检技能"""
        from agent.inspection.service import get_inspection_service

        # 优先使用 LLM 识别的意图
        intent = context.get("intent")
        if intent and intent.type == "inspection":
            info = self._convert_intent_to_info(intent)
            logger.info(f"使用 LLM 意图: {intent.action}, params: {intent.params}")
        else:
            # 降级到原有逻辑
            info = self.extract_info(text)

        action = info.get("action", "")

        if not action:
            return self._help_text()

        # 获取巡检服务
        service = get_inspection_service()

        user_id = context.get("user_id", "")
        user_nick = context.get("sender_nick", "")

        try:
            if action == "check_in":
                return await self._handle_check_in(service, info, user_id, user_nick, context)
            elif action == "check_out":
                return await self._handle_check_out(service, info, user_id, context)
            elif action == "report_issue":
                return await self._handle_report_issue(service, info, user_id, user_nick, context)
            elif action == "list_plans":
                return self._handle_list_plans(service, user_id)
            elif action == "stats":
                return self._handle_stats(service)
            elif action == "list_records":
                return self._handle_list_records(service, info, user_id, user_nick, context)
            elif action == "list_issues":
                return self._handle_list_issues(service, info, user_id, user_nick, context)
            elif action == "list_photos":
                return self._handle_list_photos(service, info, user_id, user_nick, context)
            elif action == "add_photo":
                return await self._handle_add_photo(service, info, user_id, context)
            elif action == "user_records":
                return self._handle_user_records(service, info, context)
            elif action == "point_records":
                return self._handle_point_records(service, info, context)
            elif action == "my_orders":
                return self._handle_my_orders(service, user_id, user_nick)
            elif action == "submit_rectification":
                return await self._handle_submit_rectification(service, info, user_id, context)
            elif action == "review":
                return await self._handle_review(service, info, user_id, context)
            elif action == "point_template":
                return self._handle_point_template(context)
            elif action == "issue_template":
                return self._handle_issue_template(context)
            elif action == "accept_order":
                return await self._handle_accept_order(service, info, user_id, user_nick, context)
            elif action == "complete_order":
                return await self._handle_complete_order(service, info, user_id, user_nick, context)
            else:
                return self._help_text()
        except Exception as e:
            logger.error(f"巡检技能执行失败: {e}", exc_info=True)
            return f"⚠️ 巡检操作失败：{str(e)}"

    def _convert_intent_to_info(self, intent) -> dict:
        """将 LLM 识别的意图转换为技能可用的格式"""
        info = {"action": "", "params": ""}

        action = intent.action
        params = intent.params

        # 检查 query_type 参数，区分是查询问题还是记录
        query_type = params.get("query_type", "")

        if query_type in ["问题", "问题列表", "issues", "problem", "problems"]:
            # 查询巡检问题
            info["action"] = "list_issues"
        elif query_type in ["照片", "photos", "图片"]:
            # 查询巡检照片
            info["action"] = "list_photos"
        elif action == "stats":
            # 统计
            info["action"] = "stats"
        elif action == "checkin":
            # 打卡
            info["action"] = "check_in"
        elif action == "checkout":
            # 签退
            info["action"] = "check_out"
        else:
            # 默认查询记录
            info["action"] = "list_records"

        # 构建 params
        params_parts = []

        # 时间参数
        time_str = params.get("time", "")
        if time_str:
            # 转换时间格式
            time_map = {
                "today": "今天",
                "yesterday": "昨天",
                "this_week": "这周",
                "last_week": "上周",
                "this_month": "这月",
                "last_month": "上月",
                "recent": "最近",
            }
            if time_str in time_map:
                params_parts.append(time_map[time_str])
            elif "_days_ago" in time_str:
                days = time_str.split("_")[0]
                params_parts.append(f"{days}天前")
            else:
                params_parts.append(time_str)

        # 点位名称
        point_name = params.get("point_name", "")
        if point_name:
            params_parts.append(point_name)

        # 人员名称
        person_name = params.get("person_name", "") or params.get("teacher_name", "")
        if person_name:
            params_parts.append(person_name)

        # 问题描述（用于上报问题）
        problem_desc = params.get("problem_description", "")
        if problem_desc:
            info["action"] = "report_issue"
            info["params"] = problem_desc
            return info

        # 构建 params 字符串
        info["params"] = " ".join(params_parts) if params_parts else ""

        # 如果是查询操作且没有指定时间，默认使用"最近"
        if info["action"] in ["list_records", "list_issues"] and not time_str:
            info["params"] = "最近"

        return info

    # ==================== 操作处理 ====================

    async def _handle_check_in(self, service, info, user_id, user_nick, context) -> str:
        """处理打卡签到"""
        point_name = info.get("params", "")

        if not point_name:
            # 列出可用点位
            points = service.list_points()
            if not points:
                return "📍 暂无巡检点位，请先在管理后台配置"

            lines = ["📍 可用巡检点位：\n"]
            area_names = {
                "teaching": "🏫 教学区",
                "dormitory": "🏠 宿舍区",
                "canteen": "🍽️ 食堂",
                "playground": "⚽ 操场",
                "fire": "🧯 消防设施",
                "public": "🏛️ 公共区域",
            }
            for p in points[:10]:
                area_label = area_names.get(p.area_type, p.area_type)
                lines.append(f"• {p.point_name} ({area_label})")
            lines.append("\n💡 发送「巡检打卡 点位名称」进行打卡")
            return "\n".join(lines)

        # 查找点位
        points = service.list_points()
        matched_point = None
        for p in points:
            if point_name in p.point_name or p.point_name in point_name:
                matched_point = p
                break

        if not matched_point:
            return f"❌ 未找到点位「{point_name}」\n💡 请确认点位名称，或发送「巡检打卡」查看所有点位"

        # 查找当前进行中的计划
        active_plans = service.list_plans(status="active")
        if not active_plans:
            return "❌ 当前没有进行中的巡检计划\n💡 请先在管理后台激活巡检计划"

        plan = active_plans[0]  # 使用第一个活跃计划

        # 执行打卡
        record, msg = service.check_in(
            plan_id=plan.plan_id,
            point_id=matched_point.point_id,
            inspector_id=user_id,
            inspector_name=user_nick,
        )

        if record:
            # 保存到上下文，供后续签退使用
            context["_last_checkin_record_id"] = record.record_id
            context["_file_to_send"] = None

        return msg

    async def _handle_check_out(self, service, info, user_id, context) -> str:
        """处理签退"""
        # 查找最近的打卡记录
        records = service.list_records(inspector_id=user_id)
        if not records:
            return "❌ 没有找到打卡记录\n💡 请先进行巡检打卡"

        # 找最近一条未签退的记录
        latest_record = None
        for r in records:
            if r.check_out_time == 0:
                latest_record = r
                break

        if not latest_record:
            return "✅ 你今天的所有巡检都已完成！"

        # 执行签退（默认全部通过）
        success, msg = service.check_out(
            record_id=latest_record.record_id,
            check_results=[],
        )

        if success:
            # 检查是否有问题需要上报
            if latest_record.overall_status == "has_issues":
                return f"{msg}\n💡 请发送「上报问题 [问题描述]」进行上报"
        return msg

    async def _handle_add_photo(self, service, info, user_id, context) -> str:
        """处理补充照片"""
        record_id = info.get("params", "")

        # 从上下文中获取照片
        photo_urls = context.get("_photo_urls", [])

        if not record_id:
            # 没有指定记录ID，查找最近的打卡记录
            records = service.list_records(inspector_id=user_id)
            if not records:
                return "❌ 没有找到打卡记录\n💡 请先进行巡检打卡"

            # 找最近一条记录
            record = records[0]
            record_id = record.record_id
        else:
            record = service._records.get(record_id)
            if not record:
                return f"❌ 未找到记录「{record_id}」\n💡 发送「巡检记录」查看所有记录"

        if not photo_urls:
            return (
                "💡 请发送照片来补充打卡记录\n\n"
                "使用方式：\n"
                "1. 发送「补充照片」\n"
                "2. 然后发送照片（支持多张）\n"
                "3. 系统会自动关联到最近的打卡记录"
            )

        # 保存照片并更新记录
        saved_photos = []
        for photo_url in photo_urls:
            # 下载并保存照片
            local_path = self._save_photo(photo_url, record_id)
            if local_path:
                saved_photos.append(local_path)

        if saved_photos:
            # 更新记录的照片列表
            if record.photo_urls:
                record.photo_urls.extend(saved_photos)
            else:
                record.photo_urls = saved_photos
            service._save_records()

            return (
                f"✅ 照片已补充！\n\n"
                f"📋 记录ID：{record_id}\n"
                f"📸 新增照片：{len(saved_photos)} 张\n"
                f"📷 累计照片：{len(record.photo_urls)} 张"
            )
        else:
            return "⚠️ 照片保存失败，请重试"

    def _save_photo(self, photo_url: str, record_id: str) -> Optional[str]:
        """保存照片到本地"""
        import requests

        try:
            # 创建照片目录
            os.makedirs(PHOTOS_DIR, exist_ok=True)

            # 生成文件名
            timestamp = int(time.time() * 1000)
            url_hash = hashlib.md5(photo_url.encode()).hexdigest()[:8]
            filename = f"{record_id}_{timestamp}_{url_hash}.jpg"
            filepath = os.path.join(PHOTOS_DIR, filename)

            # 下载照片
            if photo_url.startswith("http"):
                resp = requests.get(photo_url, timeout=30)
                resp.raise_for_status()
                with open(filepath, "wb") as f:
                    f.write(resp.content)
            else:
                # 本地文件路径
                import shutil
                shutil.copy2(photo_url, filepath)

            logger.info(f"保存照片成功: {filepath}")
            return filepath

        except Exception as e:
            logger.error(f"保存照片失败: {e}")
            return None

    async def _handle_report_issue(self, service, info, user_id, user_nick, context) -> str:
        """处理问题上报"""
        description = info.get("params", "")
        if not description:
            return "💡 请描述问题，例如：\n「上报问题 教学楼A走廊灯管损坏」"

        # 查找最近的打卡记录
        records = service.list_records(inspector_id=user_id)
        record = records[0] if records else None

        # 获取点位信息（用于查找负责人）
        point_name = ""
        point_id = ""
        if record:
            point_id = record.point_id
            # 从点位ID中提取名称
            for p in service.list_points():
                if p.point_id == point_id:
                    point_name = p.point_name
                    break

        # 自动分类
        category = self._auto_classify_issue(description)

        # 提取标题（取前30字）
        title = description[:30] + ("..." if len(description) > 30 else "")

        # 判断严重程度
        severity = self._judge_severity(description)

        issue = service.report_issue(
            record_id=record.record_id if record else "",
            category=category,
            title=title,
            description=description,
            reported_by=user_id,
            reported_by_name=user_nick,
            severity=severity,
            point_name=point_name,
        )

        category_names = {
            "safety_hazard": "安全隐患",
            "hygiene_issue": "卫生问题",
            "facility_damage": "设施损坏",
            "discipline_violation": "纪律违规",
            "fire_safety": "消防安全",
            "other": "其他",
        }

        # 自动派单给点位负责人
        assigned_msg = ""
        if point_id:
            points = service.list_points()
            for p in points:
                if p.point_id == point_id and p.assigned_to:
                    # 找到负责人，自动派单
                    success, msg = service.assign_order(
                        issue_id=issue.issue_id,
                        assigned_to=p.assigned_to,
                        assigned_to_name=p.assigned_to_name,
                        assigned_by=user_id,
                    )
                    if success:
                        assigned_msg = f"\n👤 已自动派单给负责人：{p.assigned_to_name}"
                    break

        return (
            f"✅ 问题已上报！\n\n"
            f"📋 问题编号：{issue.issue_id}\n"
            f"🏷️ 分类：{category_names.get(issue.category, issue.category)}\n"
            f"⚠️ 严重程度：{issue.severity}\n"
            f"📍 点位：{point_name or '未关联'}\n"
            f"📝 描述：{description}\n"
            f"{assigned_msg}\n\n"
            f"系统已自动创建整改工单。"
            f"{'负责人将收到通知并处理。' if assigned_msg else '等待管理员派单处理。'}"
        )

    def _handle_list_plans(self, service, user_id) -> str:
        """处理查看计划"""
        plans = service.list_plans()
        if not plans:
            return "📋 暂无巡检计划\n💡 请在管理后台创建巡检计划"

        status_names = {"draft": "草稿", "active": "进行中", "completed": "已完成", "cancelled": "已取消"}
        area_names = {
            "teaching": "教学区", "dormitory": "宿舍区", "canteen": "食堂",
            "playground": "操场", "fire": "消防设施", "public": "公共区域",
        }

        lines = ["📋 巡检计划列表：\n"]
        for p in plans[:5]:
            status_label = status_names.get(p.status, p.status)
            area_label = area_names.get(p.area_type, p.area_type)
            lines.append(f"• {p.plan_name}")
            lines.append(f"  区域：{area_label} | 状态：{status_label}")
            if p.start_date:
                lines.append(f"  时间：{p.start_date} ~ {p.end_date}")
            lines.append("")

        return "\n".join(lines)

    def _handle_stats(self, service) -> str:
        """处理巡检统计"""
        stats = service.get_stats()

        lines = [
            "📊 巡检统计概览：\n",
            f"📋 计划总数：{stats.total_plans}（进行中：{stats.active_plans}）",
            f"📍 巡检点位：{stats.total_points}",
            f"📝 今日打卡：{stats.total_records_today}",
            f"📈 本周打卡：{stats.total_records_week}",
            "",
            f"⚠️ 问题总数：{stats.total_issues}",
            f"  待处理：{stats.pending_issues}",
            f"  已解决：{stats.resolved_issues}",
            "",
            f"📋 工单总数：{stats.total_orders}",
            f"  待处理：{stats.pending_orders}",
        ]

        if stats.by_category:
            lines.append("\n📌 按分类：")
            category_names = {
                "safety_hazard": "安全隐患",
                "hygiene_issue": "卫生问题",
                "facility_damage": "设施损坏",
                "discipline_violation": "纪律违规",
                "fire_safety": "消防安全",
            }
            for cat, count in stats.by_category.items():
                cat_label = category_names.get(cat, cat)
                lines.append(f"  {cat_label}：{count}")

        if stats.by_severity:
            lines.append("\n🔴 按严重程度：")
            severity_names = {"low": "低", "medium": "中", "high": "高", "critical": "严重"}
            for sev, count in stats.by_severity.items():
                sev_label = severity_names.get(sev, sev)
                lines.append(f"  {sev_label}：{count}")

        return "\n".join(lines)

    def _handle_list_issues(self, service, info, user_id, user_nick, context=None) -> str:
        """处理查看巡检问题"""
        from datetime import datetime, timedelta

        params = info.get("params", "")

        # 解析时间参数
        date_str = ""
        if "最近" in params:
            # 查询最近7天的问题
            issues = service.list_issues()
            # 过滤最近7天的问题
            cutoff_time = datetime.now() - timedelta(days=7)
            issues = [i for i in issues if datetime.fromtimestamp(i.reported_at) >= cutoff_time]
            date_label = "最近7天"
        elif "今天" in params:
            today = datetime.now().strftime("%Y-%m-%d")
            issues = service.list_issues()
            issues = [i for i in issues if datetime.fromtimestamp(i.reported_at).strftime("%Y-%m-%d") == today]
            date_label = "今天"
        elif "昨天" in params:
            yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
            issues = service.list_issues()
            issues = [i for i in issues if datetime.fromtimestamp(i.reported_at).strftime("%Y-%m-%d") == yesterday]
            date_label = "昨天"
        else:
            # 默认查询最近7天
            issues = service.list_issues()
            cutoff_time = datetime.now() - timedelta(days=7)
            issues = [i for i in issues if datetime.fromtimestamp(i.reported_at) >= cutoff_time]
            date_label = "最近7天"

        if not issues:
            return f"📋 {date_label}没有巡检问题 ✅"

        # 格式化输出
        category_names = {
            "safety_hazard": "安全隐患",
            "hygiene_issue": "卫生问题",
            "facility_damage": "设施损坏",
            "discipline_violation": "纪律违规",
            "fire_safety": "消防安全",
            "other": "其他",
        }
        severity_names = {"low": "低", "medium": "中", "high": "高", "critical": "严重"}
        status_names = {
            "pending": "待处理",
            "assigned": "已派单",
            "in_progress": "整改中",
            "pending_review": "待复查",
            "resolved": "已解决",
            "closed": "已关闭",
        }

        lines = [f"⚠️ {date_label}的巡检问题：\n"]

        for issue in issues[:10]:
            report_time = datetime.fromtimestamp(issue.reported_at).strftime("%m-%d %H:%M")
            category = category_names.get(issue.category, issue.category)
            severity = severity_names.get(issue.severity, issue.severity)
            status = status_names.get(issue.status, issue.status)

            severity_icon = {"low": "🟢", "medium": "🟡", "high": "🟠", "critical": "🔴"}.get(issue.severity, "⚪")

            lines.append(f"{severity_icon} {issue.title}")
            lines.append(f"   分类：{category} | 严重程度：{severity} | 状态：{status}")
            lines.append(f"   上报时间：{report_time} | 上报人：{issue.reported_by_name}")
            lines.append(f"   问题ID：{issue.issue_id}")
            lines.append("")

        if len(issues) > 10:
            lines.append(f"... 还有 {len(issues) - 10} 个问题")

        lines.append("💡 发送「确认接收 issue_id」可接收工单处理")

        return "\n".join(lines)

    def _handle_list_photos(self, service, info, user_id, user_nick, context=None) -> str:
        """处理查看巡检照片"""
        from datetime import datetime, timedelta
        import os

        params = info.get("params", "")
        point_name = params.strip() if params else ""

        # 查找匹配的点位
        matched_point = None
        if point_name:
            for p in service.list_points():
                if point_name in p.point_name or p.point_name in point_name:
                    matched_point = p
                    break

        # 查询最近的记录
        records = service.list_records(inspector_id=user_id)
        if matched_point:
            records = [r for r in records if r.point_id == matched_point.point_id]

        # 收集照片
        all_photos = []
        for r in records:
            if r.photo_urls:
                for photo in r.photo_urls:
                    if os.path.exists(photo):
                        all_photos.append(photo)

        if not all_photos:
            point_label = f"「{matched_point.point_name}」" if matched_point else ""
            return f"📷 {point_label}没有巡检照片\n💡 打卡时可发送照片，或使用「补充照片」命令补充"

        # 设置要发送的照片
        if context is not None and all_photos:
            context["_file_to_send"] = all_photos[0]
            context["_file_type"] = "image"
            context["_file_name"] = os.path.basename(all_photos[0])

        point_label = f"「{matched_point.point_name}」" if matched_point else ""
        lines = [f"📷 {point_label}的巡检照片：\n"]
        lines.append(f"共找到 {len(all_photos)} 张照片")

        if len(all_photos) > 1:
            lines.append(f"\n💡 发送「点位记录 [点位]」可查看所有记录详情")

        if context is not None and all_photos:
            lines.append(f"\n📷 正在发送第1张照片...")

        return "\n".join(lines)

    def _handle_list_records(self, service, info, user_id, user_nick, context=None) -> str:
        """处理查看巡检记录 - 支持灵活查询

        支持格式：
        - 巡检记录                           → 今天的记录
        - 巡检记录 2026-06-22               → 指定日期的记录
        - 巡检记录 操场看台                  → 指定点位的记录
        - 巡检记录 操场看台 2026-06-22       → 指定点位和日期
        - 巡检记录 操场看台 照片             → 查看指定点位的照片
        - 巡检记录 操场看台 巡检照片         → 查看指定点位的照片
        - 操场看台 照片                      → 查看指定点位的照片
        - 操场看台 巡检照片                  → 查看指定点位的照片
        - 昨天的巡检记录                     → 自然语言日期
        - 这周操场看台的照片                 → 自然语言查询
        """
        from datetime import datetime, timedelta
        import re
        import os

        params = info.get("params", "")
        want_photos = False  # 是否想要查看照片

        # 检查是否想要查看照片
        if "照片" in params or "巡检照片" in params:
            want_photos = True
            params = params.replace("照片", "").replace("巡检照片", "").strip()

        # 解析参数：提取日期和点位名称
        date_str = ""
        point_name = ""

        logger.info(f"开始解析参数: params='{params}'")

        if params:
            # 尝试提取具体日期
            date_match = re.search(r'(\d{4}-\d{2}-\d{2})', params)
            if date_match:
                date_str = date_match.group(1)
                # 剩余部分作为点位名称
                remaining = params.replace(date_str, "").strip()
                if remaining:
                    point_name = remaining
                logger.info(f"解析到具体日期: {date_str}")
            else:
                # 尝试解析自然语言日期
                logger.info(f"尝试解析自然语言日期: '{params}'")
                date_str = self._parse_natural_date(params)
                logger.info(f"自然语言日期解析结果: '{date_str}'")
                if date_str:
                    # 移除日期描述，剩余作为点位名称
                    for keyword in ["昨天", "今天", "前天", "本周", "这周", "上周"]:
                        params = params.replace(keyword, "")
                    point_name = params.strip()
                    logger.info(f"解析到自然语言日期: {date_str}, 点位: '{point_name}'")
                else:
                    # 没有日期，整个作为点位名称
                    point_name = params.strip()
                    logger.info(f"未解析到日期，作为点位名称: '{point_name}'")

        # 如果没有指定日期，使用今天
        if not date_str:
            today = datetime.now().strftime("%Y-%m-%d")
            date_str = today
            date_label = f"今天 ({today})"
        else:
            date_label = date_str

        # 查找匹配的点位
        matched_point = None
        if point_name:
            for p in service.list_points():
                if point_name in p.point_name or p.point_name in point_name:
                    matched_point = p
                    break

        # 查询记录
        records = []
        if matched_point:
            # 查询指定点位的记录
            all_records = service.list_records(inspector_id=user_id, date_str=date_str)
            for r in all_records:
                if r.point_id == matched_point.point_id:
                    records.append(r)
            point_label = f"「{matched_point.point_name}」"
        else:
            # 查询所有记录
            records = service.list_records(inspector_id=user_id, date_str=date_str)
            point_label = ""

        # 如果是"最近"查询，需要查询多天的记录
        if "最近" in params and not records:
            # 查询最近7天的记录
            from datetime import datetime, timedelta
            today = datetime.now()
            for i in range(7):
                date = today - timedelta(days=i)
                date_str = date.strftime("%Y-%m-%d")
                daily_records = service.list_records(inspector_id=user_id, date_str=date_str)
                records.extend(daily_records)
            date_label = "最近7天"

        if not records:
            return f"📋 {user_nick}，{date_label} {point_label}没有巡检记录\n💡 发送「巡检打卡 [点位名称]」开始巡检"

        # 如果用户想要查看照片
        if want_photos:
            return self._handle_list_record_photos(records, service, point_label, date_label, context)

        # 显示记录列表
        lines = [f"📋 {user_nick} 的巡检记录 ({date_label}{point_label})：\n"]

        for r in records[:10]:
            check_in_time = datetime.fromtimestamp(r.check_in_time).strftime("%H:%M")
            check_out_time = datetime.fromtimestamp(r.check_out_time).strftime("%H:%M") if r.check_out_time else "未签退"

            # 获取点位名称
            point = service.get_point(r.point_id)
            p_name = point.point_name if point else "未知点位"

            status_icon = "✅" if r.overall_status == "normal" else "⚠️"
            photo_count = len(r.photo_urls) if r.photo_urls else 0
            photo_icon = f" 📸×{photo_count}" if photo_count > 0 else ""

            lines.append(f"{status_icon} {p_name}{photo_icon}")
            lines.append(f"   打卡：{check_in_time} | 签退：{check_out_time}")
            lines.append(f"   记录ID：{r.record_id}")

        if len(records) > 10:
            lines.append(f"\n... 还有 {len(records) - 10} 条记录")

        lines.append("\n💡 发送「巡检记录 [点位] 照片」可查看打卡照片")

        return "\n".join(lines)

    def _handle_list_record_photos(self, records, service, point_label, date_label, context=None) -> str:
        """处理查看巡检记录照片"""
        import os

        # 收集所有照片
        all_photos = []
        for r in records:
            if r.photo_urls:
                for photo in r.photo_urls:
                    if os.path.exists(photo):
                        all_photos.append(photo)

        if not all_photos:
            return f"📷 {date_label}{point_label} 没有打卡照片\n💡 打卡时可发送照片，或使用「补充照片」命令补充"

        # 设置要发送的照片
        if context is not None:
            # 发送第一张照片
            context["_file_to_send"] = all_photos[0]
            context["_file_type"] = "image"
            context["_file_name"] = os.path.basename(all_photos[0])

        lines = [f"📷 {date_label}{point_label} 的打卡照片：\n"]
        lines.append(f"共找到 {len(all_photos)} 张照片")

        if len(all_photos) > 1:
            lines.append(f"\n💡 发送「点位记录 [点位]」可查看所有记录详情")

        if context is not None and len(all_photos) > 0:
            lines.append(f"\n📷 正在发送第1张照片...")

        return "\n".join(lines)

    def _handle_user_records(self, service, info, context) -> str:
        """处理查看指定用户的巡检记录

        支持格式：
        - 用户记录 启拓
        - 用户记录 启拓 操场看台
        - 用户记录 启拓 2026-06-16
        - 用户记录 启拓 操场看台 2026-06-16
        """
        from datetime import datetime
        import os

        params = info.get("params", "")
        if not params:
            return (
                "💡 请指定查询条件，例如：\n"
                "「用户记录 启拓」- 查看启拓的所有记录\n"
                "「用户记录 启拓 操场看台」- 查看启拓在操场看台的记录\n"
                "「用户记录 启拓 2026-06-16」- 查看启拓今天的记录\n"
                "「用户记录 启拓 操场看台 2026-06-16」- 查看启拓在操场看台今天的记录"
            )

        # 解析参数
        parts = params.split()
        user_name = parts[0] if len(parts) >= 1 else ""
        point_name = ""
        date_str = ""

        # 检测日期格式 (YYYY-MM-DD)
        for part in parts[1:]:
            if len(part) == 10 and part[4] == '-' and part[7] == '-':
                date_str = part
            elif not point_name:
                point_name = part

        # 搜索所有记录
        all_records = service.list_records()
        matched_records = []

        for r in all_records:
            # 匹配用户
            if user_name not in (r.inspector_name or ""):
                continue

            # 匹配点位
            if point_name:
                point = service.get_point(r.point_id)
                if point and point_name not in point.point_name:
                    continue

            # 匹配日期
            if date_str:
                record_date = datetime.fromtimestamp(r.check_in_time).strftime("%Y-%m-%d")
                if record_date != date_str:
                    continue

            matched_records.append(r)

        if not matched_records:
            return f"❌ 未找到匹配的巡检记录"

        # 构建标题
        title_parts = [f"📋 {user_name}"]
        if point_name:
            title_parts.append(f"在「{point_name}」")
        if date_str:
            title_parts.append(f"({date_str})")
        title_parts.append("的巡检记录：")
        lines = [" ".join(title_parts) + "\n"]

        # 收集所有照片
        all_photos = []

        for r in matched_records[:10]:
            check_in_time = datetime.fromtimestamp(r.check_in_time).strftime("%Y-%m-%d %H:%M")
            check_out_time = datetime.fromtimestamp(r.check_out_time).strftime("%H:%M") if r.check_out_time else "未签退"

            # 获取点位名称
            point = service.get_point(r.point_id)
            p_name = point.point_name if point else "未知点位"

            status_icon = "✅" if r.overall_status == "normal" else "⚠️"
            photo_count = len(r.photo_urls) if r.photo_urls else 0
            photo_icon = f" 📸×{photo_count}" if photo_count > 0 else ""

            lines.append(f"{status_icon} {p_name}{photo_icon}")
            lines.append(f"   时间：{check_in_time} ~ {check_out_time}")
            lines.append(f"   记录ID：{r.record_id}")

            # 收集照片
            if r.photo_urls:
                for photo in r.photo_urls:
                    if os.path.exists(photo):
                        all_photos.append(photo)

        if len(matched_records) > 10:
            lines.append(f"\n... 还有 {len(matched_records) - 10} 条记录")

        # 设置待发送的照片
        if all_photos:
            context["_file_to_send"] = all_photos[0]  # 发送第一张照片
            context["_file_type"] = "image"
            context["_file_name"] = os.path.basename(all_photos[0])
            lines.append(f"\n📷 共找到 {len(all_photos)} 张照片，正在发送...")
        else:
            lines.append(f"\n💡 该用户暂无打卡照片")

        lines.append(f"\n💡 发送「补充照片 [记录ID]」可为记录补充照片")

        return "\n".join(lines)

    def _handle_point_records(self, service, info, context) -> str:
        """处理查看指定点位的巡检记录

        支持格式：
        - 点位记录 操场看台
        - 点位记录 操场看台 2026-06-16
        - 点位记录 操场看台 启拓
        """
        from datetime import datetime
        import os

        params = info.get("params", "")
        if not params:
            return (
                "💡 请指定点位名称，例如：\n"
                "「点位记录 操场看台」- 查看操场看台的所有记录\n"
                "「点位记录 操场看台 2026-06-16」- 查看操场看台今天的记录\n"
                "「点位记录 操场看台 启拓」- 查看启拓在操场看台的记录"
            )

        # 解析参数
        parts = params.split()
        point_name = parts[0] if len(parts) >= 1 else ""
        user_name = ""
        date_str = ""

        # 检测日期格式 (YYYY-MM-DD) 和用户
        for part in parts[1:]:
            if len(part) == 10 and part[4] == '-' and part[7] == '-':
                date_str = part
            elif not user_name:
                user_name = part

        # 查找点位
        matched_point = None
        for p in service.list_points():
            if point_name in p.point_name or p.point_name in point_name:
                matched_point = p
                break

        if not matched_point:
            return f"❌ 未找到点位「{point_name}」"

        # 如果没有指定日期，默认查询今天的记录
        if not date_str:
            today = datetime.now().strftime("%Y-%m-%d")
            date_str = today

        # 搜索所有记录
        all_records = service.list_records()
        matched_records = []

        for r in all_records:
            # 匹配点位
            if r.point_id != matched_point.point_id:
                continue

            # 匹配用户
            if user_name and user_name not in (r.inspector_name or ""):
                continue

            # 匹配日期
            record_date = datetime.fromtimestamp(r.check_in_time).strftime("%Y-%m-%d")
            if record_date != date_str:
                continue

            matched_records.append(r)

        if not matched_records:
            return f"📍 {matched_point.point_name} 在 {date_str} 没有巡检记录"

        # 构建标题
        title_parts = [f"📍 {matched_point.point_name}"]
        if user_name:
            title_parts.append(f"({user_name}巡检)")
        title_parts.append(f"({date_str}) 的记录：")
        lines = [" ".join(title_parts) + "\n"]

        # 收集所有照片
        all_photos = []

        for r in matched_records[:10]:
            check_in_time = datetime.fromtimestamp(r.check_in_time).strftime("%H:%M")
            check_out_time = datetime.fromtimestamp(r.check_out_time).strftime("%H:%M") if r.check_out_time else "未签退"

            status_icon = "✅" if r.overall_status == "normal" else "⚠️"
            photo_count = len(r.photo_urls) if r.photo_urls else 0
            photo_icon = f" 📸×{photo_count}" if photo_count > 0 else ""

            lines.append(f"{status_icon} {r.inspector_name}{photo_icon}")
            lines.append(f"   时间：{check_in_time} ~ {check_out_time}")
            lines.append(f"   记录ID：{r.record_id}")

            # 收集照片
            if r.photo_urls:
                for photo in r.photo_urls:
                    if os.path.exists(photo):
                        all_photos.append(photo)

        if len(matched_records) > 10:
            lines.append(f"\n... 还有 {len(matched_records) - 10} 条记录")

        # 设置待发送的照片
        if all_photos:
            context["_file_to_send"] = all_photos[0]
            context["_file_type"] = "image"
            context["_file_name"] = os.path.basename(all_photos[0])
            lines.append(f"\n📷 共找到 {len(all_photos)} 张照片，正在发送...")
        else:
            lines.append(f"\n💡 该点位暂无打卡照片")

        return "\n".join(lines)

    def _handle_my_orders(self, service, user_id, user_nick) -> str:
        """处理查看我的工单"""
        orders = service.get_orders(assigned_to=user_id)

        if not orders:
            return f"📋 {user_nick}，你当前没有待处理的工单 ✅"

        status_names = {
            "pending": "待处理", "assigned": "已派单", "in_progress": "整改中",
            "pending_review": "待复查", "resolved": "已解决", "closed": "已关闭",
        }

        lines = [f"📋 {user_nick} 的工单：\n"]
        for o in orders[:5]:
            issue = service.get_issue(o.issue_id)
            status_label = status_names.get(o.status, o.status)
            title = issue.title if issue else "未知问题"
            lines.append(f"• [{status_label}] {title}")
            lines.append(f"  编号：{o.issue_id}")

        return "\n".join(lines)

    async def _handle_submit_rectification(self, service, info, user_id, context) -> str:
        """处理提交整改"""
        issue_id = info.get("params", "")
        if not issue_id:
            return "💡 请提供问题编号，例如：\n「整改完成 issue_xxx」"

        success, msg = service.submit_rectification(
            issue_id=issue_id,
            operator=user_id,
        )
        return msg

    async def _handle_review(self, service, info, user_id, context) -> str:
        """处理复查"""
        params = info.get("params", {})
        if not isinstance(params, dict) or "issue_id" not in params:
            return "💡 请提供问题编号，例如：\n「复查 issue_xxx 通过」\n「复查 issue_xxx 不通过」"

        success, msg = service.review_issue(
            issue_id=params["issue_id"],
            review_result=params.get("result", "pass"),
            reviewer_id=user_id,
        )
        return msg

    async def _handle_accept_order(self, service, info, user_id, user_nick, context) -> str:
        """处理负责人确认接收工单"""
        issue_id = info.get("params", "")
        if not issue_id:
            # 查找分配给该用户但未确认的工单
            orders = service.get_orders(assigned_to=user_id, status="assigned")
            if not orders:
                return f"📋 {user_nick}，你当前没有待确认的工单"

            lines = [f"📋 {user_nick}，以下工单等待你确认接收：\n"]
            for o in orders[:5]:
                issue = service.get_issue(o.issue_id)
                title = issue.title if issue else "未知问题"
                lines.append(f"• {title}")
                lines.append(f"  编号：{o.issue_id}")
            lines.append("\n💡 发送「确认接收 工单编号」确认接收")
            return "\n".join(lines)

        # 确认接收工单
        issue = service.get_issue(issue_id)
        if not issue:
            return f"❌ 工单 {issue_id} 不存在"

        if issue.assigned_to != user_id:
            return f"❌ 该工单未分配给你，无法确认接收"

        if issue.status != "assigned":
            return f"❌ 该工单当前状态为「{issue.status}」，无法确认接收"

        # 更新工单状态为整改中
        success, msg = service.start_rectification(
            issue_id=issue_id,
            operator=user_id,
        )

        if success:
            return (
                f"✅ 工单已确认接收！\n\n"
                f"📋 工单编号：{issue_id}\n"
                f"📝 问题：{issue.title}\n"
                f"⏰ 开始整改时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n"
                f"💡 维修完成后，请发送「工单完成 {issue_id}」确认完成"
            )
        else:
            return f"❌ 确认接收失败：{msg}"

    async def _handle_complete_order(self, service, info, user_id, user_nick, context) -> str:
        """处理工单完成确认"""
        issue_id = info.get("params", "")
        if not issue_id:
            # 查找该用户正在进行中的工单
            orders = service.get_orders(assigned_to=user_id, status="in_progress")
            if not orders:
                return f"📋 {user_nick}，你当前没有进行中的工单"

            lines = [f"📋 {user_nick}，以下工单正在进行中：\n"]
            for o in orders[:5]:
                issue = service.get_issue(o.issue_id)
                title = issue.title if issue else "未知问题"
                lines.append(f"• {title}")
                lines.append(f"  编号：{o.issue_id}")
            lines.append("\n💡 发送「工单完成 工单编号」确认维修完成")
            return "\n".join(lines)

        # 确认工单完成
        issue = service.get_issue(issue_id)
        if not issue:
            return f"❌ 工单 {issue_id} 不存在"

        if issue.assigned_to != user_id:
            return f"❌ 该工单未分配给你，无法确认完成"

        if issue.status != "in_progress":
            return f"❌ 该工单当前状态为「{issue.status}」，无法确认完成"

        # 提交整改结果
        success, msg = service.submit_rectification(
            issue_id=issue_id,
            rectification_notes=f"由负责人 {user_nick} 确认维修完成",
            operator=user_id,
        )

        if success:
            return (
                f"✅ 工单已确认完成！\n\n"
                f"📋 工单编号：{issue_id}\n"
                f"📝 问题：{issue.title}\n"
                f"👤 负责人：{user_nick}\n"
                f"⏰ 完成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n"
                f"系统将安排复查验收。"
            )
        else:
            return f"❌ 确认完成失败：{msg}"

    def _handle_point_template(self, context) -> str:
        """生成巡检点位导入模板"""
        output_dir = os.path.join(os.path.dirname(__file__), "..", "..", "test_output")
        output_dir = os.path.abspath(output_dir)
        os.makedirs(output_dir, exist_ok=True)

        if HAS_OPENPYXL:
            filepath = os.path.join(output_dir, "巡检点位导入模板.xlsx")
            self._create_point_excel_template(filepath)
            filename = "巡检点位导入模板.xlsx"
        else:
            filepath = os.path.join(output_dir, "巡检点位导入模板.csv")
            self._create_point_csv_template(filepath)
            filename = "巡检点位导入模板.csv"

        context["_file_to_send"] = filepath
        context["_file_name"] = filename
        context["_file_type"] = "file"

        return (
            "📄 巡检点位导入模板已生成！\n\n"
            "📋 模板字段说明：\n"
            "• 点位名称：必填\n"
            "• 区域类型：必填 teaching/dormitory/canteen/playground/fire/public\n"
            "• 位置描述：必填\n"
            "• 纬度/经度：选填，用于定位验证\n"
            "• 是否要求拍照：是/否\n"
            "• 是否要求定位：是/否\n"
            "• 检查人（负责老师）：必填\n"
            "• 检查时间：必填，如 每天 08:00、每周一 09:00\n"
            "• 检查项：选填，多个用逗号分隔\n\n"
            "💡 填写完成后，将文件发送给我即可批量导入点位。"
        )

    def _create_point_excel_template(self, filepath: str):
        """创建 Excel 格式的点位模板"""
        wb = Workbook()
        ws = wb.active
        ws.title = "巡检点位"

        # 表头样式
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        # 表头
        headers = ["点位名称", "区域类型", "位置描述", "纬度", "经度", "是否要求拍照", "是否要求定位", "检查人（负责老师）", "检查时间", "检查项（逗号分隔）"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # 示例数据
        examples = [
            ["教学楼A-1楼走廊", "teaching", "教学楼A栋1层走廊", 30.5728, 104.0668, "是", "否", "张老师", "每天 08:00", "照明设施,卫生状况,灭火器"],
            ["学生宿舍B栋入口", "dormitory", "宿舍B栋入口处", "", "", "是", "是", "李老师", "每天 22:00", "门禁系统,走廊卫生"],
            ["食堂操作间", "canteen", "食堂2楼操作间", "", "", "是", "否", "王主管", "每天 10:00", "食品安全,卫生清洁,燃气设备"],
            ["操场看台", "playground", "操场东侧看台", "", "", "是", "否", "赵老师", "每周一 09:00", "座椅安全,地面清洁"],
            ["消防通道A", "fire", "教学楼A栋消防通道", "", "", "是", "是", "刘主任", "每天 09:00", "通道畅通,应急灯,灭火器"],
        ]

        example_font = Font(color="808080", italic=True)
        for row_idx, row_data in enumerate(examples, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = example_font
                cell.border = thin_border

        # 设置列宽
        col_widths = [20, 15, 25, 12, 12, 15, 15, 18, 18, 30]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        # 添加说明 Sheet
        ws2 = wb.create_sheet("填写说明")
        ws2["A1"] = "区域类型可选值"
        ws2["A1"].font = Font(bold=True, size=12)
        area_types = [
            ("teaching", "教学区"),
            ("dormitory", "宿舍区"),
            ("canteen", "食堂"),
            ("playground", "操场"),
            ("fire", "消防设施"),
            ("public", "公共区域"),
        ]
        for i, (code, name) in enumerate(area_types, 3):
            ws2.cell(row=i, column=1, value=code)
            ws2.cell(row=i, column=2, value=name)

        ws2.cell(row=11, column=1, value="检查时间格式说明").font = Font(bold=True, size=12)
        ws2.cell(row=12, column=1, value="每天 HH:MM")
        ws2.cell(row=13, column=1, value="每周一 HH:MM")
        ws2.cell(row=14, column=1, value="自定义描述")

        ws2.column_dimensions["A"].width = 20
        ws2.column_dimensions["B"].width = 15

        wb.save(filepath)

    def _create_point_csv_template(self, filepath: str):
        """创建 CSV 格式的点位模板（备用）"""
        with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["点位名称", "区域类型", "位置描述", "纬度", "经度", "是否要求拍照", "是否要求定位", "检查项（逗号分隔）"])
            writer.writerow(["教学楼A-1楼走廊", "teaching", "教学楼A栋1层走廊", "30.5728", "104.0668", "是", "否", "照明设施,卫生状况,灭火器"])
            writer.writerow(["学生宿舍B栋入口", "dormitory", "宿舍B栋入口处", "", "", "是", "是", "门禁系统,走廊卫生"])
            writer.writerow(["食堂操作间", "canteen", "食堂2楼操作间", "", "", "是", "否", "食品安全,卫生清洁,燃气设备"])

    def _handle_issue_template(self, context) -> str:
        """生成问题上报模板"""
        output_dir = os.path.join(os.path.dirname(__file__), "..", "..", "test_output")
        output_dir = os.path.abspath(output_dir)
        os.makedirs(output_dir, exist_ok=True)

        if HAS_OPENPYXL:
            filepath = os.path.join(output_dir, "巡检问题上报模板.xlsx")
            self._create_issue_excel_template(filepath)
            filename = "巡检问题上报模板.xlsx"
        else:
            filepath = os.path.join(output_dir, "巡检问题上报模板.csv")
            self._create_issue_csv_template(filepath)
            filename = "巡检问题上报模板.csv"

        context["_file_to_send"] = filepath
        context["_file_name"] = filename
        context["_file_type"] = "file"

        return (
            "📄 巡检问题上报模板已生成！\n\n"
            "📋 模板字段说明：\n"
            "• 问题标题：必填\n"
            "• 问题分类：safety_hazard/hygiene_issue/facility_damage/discipline_violation/fire_safety/other\n"
            "• 严重程度：low/medium/high/critical\n"
            "• 问题描述：详细描述\n"
            "• 点位名称：关联的巡检点位\n"
            "• 区域类型：teaching/dormitory/canteen/playground/fire/public\n"
            "• 上报人：发现人姓名\n"
            "• 上报时间：发现问题的时间\n\n"
            "💡 填写完成后，将文件发送给我即可批量上报问题。"
        )

    def _create_issue_excel_template(self, filepath: str):
        """创建 Excel 格式的问题模板"""
        wb = Workbook()
        ws = wb.active
        ws.title = "巡检问题"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        headers = ["问题标题", "问题分类", "严重程度", "问题描述", "点位名称", "区域类型", "上报人", "上报时间"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        examples = [
            ["教学楼A走廊灯管损坏", "facility_damage", "medium", "1楼走廊东侧第3个灯管不亮", "教学楼A-1楼走廊", "teaching", "张老师", "2026-06-15 09:30"],
            ["食堂操作间地面积水", "hygiene_issue", "high", "操作间入口处地面有大面积积水", "食堂操作间", "canteen", "王主管", "2026-06-15 10:15"],
            ["宿舍B栋门禁失灵", "safety_hazard", "critical", "门禁系统无法正常识别", "学生宿舍B栋入口", "dormitory", "李老师", "2026-06-15 22:00"],
            ["操场跑道破损", "facility_damage", "medium", "跑道有多处裂缝", "操场跑道", "playground", "赵老师", "2026-06-15 08:45"],
        ]

        example_font = Font(color="808080", italic=True)
        for row_idx, row_data in enumerate(examples, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = example_font
                cell.border = thin_border

        col_widths = [25, 20, 12, 35, 20, 15, 15, 20]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        # 填写说明
        ws2 = wb.create_sheet("填写说明")
        ws2["A1"] = "问题分类可选值"
        ws2["A1"].font = Font(bold=True, size=12)
        categories = [
            ("safety_hazard", "安全隐患"),
            ("hygiene_issue", "卫生问题"),
            ("facility_damage", "设施损坏"),
            ("discipline_violation", "纪律违规"),
            ("fire_safety", "消防安全"),
            ("other", "其他"),
        ]
        for i, (code, name) in enumerate(categories, 3):
            ws2.cell(row=i, column=1, value=code)
            ws2.cell(row=i, column=2, value=name)

        ws2["A10"] = "严重程度可选值"
        ws2["A10"].font = Font(bold=True, size=12)
        severities = [("low", "低"), ("medium", "中"), ("high", "高"), ("critical", "严重")]
        for i, (code, name) in enumerate(severities, 12):
            ws2.cell(row=i, column=1, value=code)
            ws2.cell(row=i, column=2, value=name)

        ws2.column_dimensions["A"].width = 20
        ws2.column_dimensions["B"].width = 15

        wb.save(filepath)

    def _create_issue_csv_template(self, filepath: str):
        """创建 CSV 格式的问题模板（备用）"""
        with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["问题标题", "问题分类", "严重程度", "问题描述", "点位名称", "区域类型"])
            writer.writerow(["教学楼A走廊灯管损坏", "facility_damage", "medium", "1楼走廊东侧第3个灯管不亮", "教学楼A-1楼走廊", "teaching"])
            writer.writerow(["食堂操作间地面积水", "hygiene_issue", "high", "操作间入口处地面有大面积积水", "食堂操作间", "canteen"])

    # ==================== 工具方法 ====================

    def _auto_classify_issue(self, description: str) -> str:
        """自动分类问题"""
        desc_lower = description.lower()

        # 安全隐患
        safety_kw = ["危险", "安全", "隐患", "滑倒", "摔伤", "触电", "漏水", "漏电", "破损", "裂缝"]
        if any(kw in desc_lower for kw in safety_kw):
            return "safety_hazard"

        # 消防安全
        fire_kw = ["消防", "灭火器", "烟雾", "火灾", "疏散", "逃生", "消防通道", "应急灯"]
        if any(kw in desc_lower for kw in fire_kw):
            return "fire_safety"

        # 卫生问题
        hygiene_kw = ["卫生", "脏", "乱", "垃圾", "清洁", "异味", "虫", "鼠", "蟑螂", "污渍"]
        if any(kw in desc_lower for kw in hygiene_kw):
            return "hygiene_issue"

        # 设施损坏
        facility_kw = ["损坏", "故障", "坏了", "修", "灯", "门", "窗", "桌椅", "设备", "空调"]
        if any(kw in desc_lower for kw in facility_kw):
            return "facility_damage"

        # 纪律违规
        discipline_kw = ["纪律", "违规", "吵闹", "打架", "迟到", "早退", "旷课", "抽烟"]
        if any(kw in desc_lower for kw in discipline_kw):
            return "discipline_violation"

        return "other"

    def _judge_severity(self, description: str) -> str:
        """判断问题严重程度"""
        desc_lower = description.lower()

        # 严重
        critical_kw = ["火灾", "触电", "重伤", "坍塌", "中毒", "危险"]
        if any(kw in desc_lower for kw in critical_kw):
            return "critical"

        # 高
        high_kw = ["漏水", "漏电", "损坏", "故障", "松动", "脱落", "裂缝"]
        if any(kw in desc_lower for kw in high_kw):
            return "high"

        # 低
        low_kw = ["轻微", "少许", "一点", "小问题"]
        if any(kw in desc_lower for kw in low_kw):
            return "low"

        return "medium"

    def _parse_natural_date(self, text: str) -> str:
        """解析自然语言日期，返回YYYY-MM-DD格式"""
        from datetime import datetime, timedelta

        today = datetime.now()
        text_lower = text.lower()

        # 最近（返回7天前的日期）- 支持中英文
        if "最近" in text_lower or "recent" in text_lower:
            week_ago = today - timedelta(days=7)
            return week_ago.strftime("%Y-%m-%d")

        # 今天
        if "今天" in text_lower or "今日" in text_lower:
            return today.strftime("%Y-%m-%d")

        # 昨天
        if "昨天" in text_lower or "昨日" in text_lower:
            yesterday = today - timedelta(days=1)
            return yesterday.strftime("%Y-%m-%d")

        # 前天
        if "前天" in text_lower:
            day_before = today - timedelta(days=2)
            return day_before.strftime("%Y-%m-%d")

        # 大前天
        if "大前天" in text_lower:
            day_before = today - timedelta(days=3)
            return day_before.strftime("%Y-%m-%d")

        # 明天
        if "明天" in text_lower or "明日" in text_lower:
            tomorrow = today + timedelta(days=1)
            return tomorrow.strftime("%Y-%m-%d")

        # 后天
        if "后天" in text_lower:
            day_after = today + timedelta(days=2)
            return day_after.strftime("%Y-%m-%d")

        # 本周/这周
        if "本周" in text_lower or "这周" in text_lower:
            # 返回本周一的日期
            weekday = today.weekday()
            monday = today - timedelta(days=weekday)
            return monday.strftime("%Y-%m-%d")

        # 上周
        if "上周" in text_lower:
            weekday = today.weekday()
            last_monday = today - timedelta(days=weekday + 7)
            return last_monday.strftime("%Y-%m-%d")

        # 下周
        if "下周" in text_lower:
            weekday = today.weekday()
            next_monday = today + timedelta(days=7 - weekday)
            return next_monday.strftime("%Y-%m-%d")

        # 周一到周日
        weekday_map = {
            '周一': 0, '星期一': 0,
            '周二': 1, '星期二': 1,
            '周三': 2, '星期三': 2,
            '周四': 3, '星期四': 3,
            '周五': 4, '星期五': 4,
            '周六': 5, '星期六': 5,
            '周日': 6, '星期日': 6, '周天': 6, '星期天': 6,
        }
        for day_name, weekday_num in weekday_map.items():
            if day_name in text_lower:
                # 计算最近的这个周几
                current_weekday = today.weekday()
                days_diff = weekday_num - current_weekday
                if days_diff <= 0:
                    # 如果目标日期已过，取下周的
                    days_diff += 7
                target_date = today + timedelta(days=days_diff)
                return target_date.strftime("%Y-%m-%d")

        # X天前
        import re
        days_ago_match = re.search(r'(\d+)\s*天前', text_lower)
        if days_ago_match:
            days = int(days_ago_match.group(1))
            target_date = today - timedelta(days=days)
            return target_date.strftime("%Y-%m-%d")

        # X天后
        days_later_match = re.search(r'(\d+)\s*天后', text_lower)
        if days_later_match:
            days = int(days_later_match.group(1))
            target_date = today + timedelta(days=days)
            return target_date.strftime("%Y-%m-%d")

        # 无法解析
        return ""

    def _help_text(self) -> str:
        """帮助文本"""
        return (
            "🔍 巡检管理指令：\n\n"
            "📍 巡检打卡：\n"
            "  「巡检打卡」- 查看所有点位\n"
            "  「巡检打卡 教学楼A」- 在指定点位打卡\n\n"
            "📸 照片管理：\n"
            "  「补充照片」- 为最近打卡记录补充照片\n"
            "  「补充照片 record_xxx」- 为指定记录补充照片\n\n"
            "📝 问题上报：\n"
            "  「上报问题 教学楼走廊灯管损坏」\n\n"
            "📋 计划与记录：\n"
            "  「巡检计划」- 查看所有计划\n"
            "  「巡检统计」- 查看统计数据\n"
            "  「巡检记录」- 查看今天的巡检记录\n"
            "  「巡检记录 2026-06-16」- 查看指定日期记录\n"
            "  「巡检记录 操场看台」- 查看指定点位的记录\n"
            "  「巡检记录 操场看台 2026-06-16」- 查看指定点位和日期的记录\n"
            "  「巡检记录 操场看台 照片」- 查看指定点位的打卡照片\n\n"
            "👤 用户记录查询：\n"
            "  「用户记录 启拓」- 查看启拓的所有记录\n"
            "  「用户记录 启拓 操场看台」- 查看启拓在操场看台的记录\n"
            "  「用户记录 启拓 2026-06-16」- 查看启拓今天的记录\n\n"
            "📍 点位记录查询：\n"
            "  「点位记录 操场看台」- 查看操场看台的所有记录\n"
            "  「点位记录 操场看台 2026-06-16」- 查看操场看台今天的记录\n"
            "  「点位记录 操场看台 启拓」- 查看启拓在操场看台的记录\n\n"
            "📷 点位照片查询：\n"
            "  「操场看台 照片」- 查看操场看台的打卡照片\n"
            "  「操场看台 巡检照片」- 查看操场看台的巡检照片\n\n"
            "📋 工单处理（负责人）：\n"
            "  「工单列表」- 查看我的工单\n"
            "  「确认接收」- 查看待确认的工单\n"
            "  「确认接收 issue_xxx」- 确认接收工单\n"
            "  「工单完成」- 查看进行中的工单\n"
            "  「工单完成 issue_xxx」- 确认维修完成\n"
            "  「整改完成 issue_xxx」- 提交整改\n"
            "  「复查 issue_xxx 通过」- 复查验收\n\n"
            "📄 模板下载：\n"
            "  「巡检点位模板」- 下载点位导入模板\n"
            "  「巡检问题模板」- 下载问题上报模板\n\n"
            "🔄 签退：\n"
            "  「巡检签退」- 完成本次巡检"
        )


# 注册技能
skill_registry.register(InspectionSkill())
