"""
Excel 数据处理模块

处理排课数据的导入导出：
- 生成 Excel 模板
- 解析 Excel 数据
- 导出课表到 Excel
"""
import os
import logging
from typing import Dict, List, Optional

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("openpyxl 未安装，Excel 功能不可用。请运行: pip install openpyxl")


def generate_template(output_path: str) -> bool:
    """
    生成排课数据 Excel 模板

    参数:
        output_path: 输出文件路径

    返回:
        是否成功
    """
    if not HAS_OPENPYXL:
        logger.error("openpyxl 未安装")
        return False

    try:
        wb = openpyxl.Workbook()

        # 样式定义
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=12, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        def style_header(ws, headers, row=1):
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

        # ========== 班级信息表 ==========
        ws_classes = wb.active
        ws_classes.title = "班级信息"
        class_headers = ["班级ID", "班级名称", "年级", "学生人数", "班主任", "课程ID列表(逗号分隔)", "固定教室(逗号分隔)"]
        style_header(ws_classes, class_headers)

        # 示例数据
        ws_classes.append(["class_01", "高一(1)班", "高一", 45, "张老师", "math,chinese,english", "101教室,102教室"])
        ws_classes.append(["class_02", "高一(2)班", "高一", 46, "李老师", "math,chinese,english", "103教室,104教室"])

        # 设置列宽
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws_classes.column_dimensions[col_letter].width = 18

        # ========== 教师信息表 ==========
        ws_teachers = wb.create_sheet("教师信息")
        teacher_headers = ["教师ID", "教师姓名", "可教科目(逗号分隔)", "每天最大课时", "每周最大课时"]
        style_header(ws_teachers, teacher_headers)

        ws_teachers.append(["teacher_01", "张老师", "数学", 4, 20])
        ws_teachers.append(["teacher_02", "李老师", "语文", 4, 20])
        ws_teachers.append(["teacher_03", "王老师", "英语", 4, 20])

        for col_letter in ['A', 'B', 'C', 'D', 'E']:
            ws_teachers.column_dimensions[col_letter].width = 18

        # ========== 课程信息表 ==========
        ws_courses = wb.create_sheet("课程信息")
        course_headers = ["课程ID", "课程名称", "学科", "每周课时数", "是否主课", "需要连排", "所需设备(逗号分隔)"]
        style_header(ws_courses, course_headers)

        ws_courses.append(["math", "高一数学", "数学", 5, "是", "否", ""])
        ws_courses.append(["chinese", "高一语文", "语文", 5, "是", "否", ""])
        ws_courses.append(["english", "高一英语", "英语", 5, "是", "否", ""])
        ws_courses.append(["physics", "高一物理", "物理", 3, "否", "否", "实验室"])

        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws_courses.column_dimensions[col_letter].width = 18

        # ========== 教室信息表 ==========
        ws_classrooms = wb.create_sheet("教室信息")
        classroom_headers = ["教室ID", "教室名称", "容量", "设备(逗号分隔)", "教学楼"]
        style_header(ws_classrooms, classroom_headers)

        ws_classrooms.append(["room_01", "101教室", 50, "多媒体", "教学楼A"])
        ws_classrooms.append(["room_02", "102教室", 50, "多媒体", "教学楼A"])
        ws_classrooms.append(["room_03", "实验室1", 45, "实验设备,多媒体", "实验楼"])

        for col_letter in ['A', 'B', 'C', 'D', 'E']:
            ws_classrooms.column_dimensions[col_letter].width = 18

        # ========== 约束配置表 ==========
        ws_constraints = wb.create_sheet("约束配置")
        constraint_headers = ["约束项", "值", "说明"]
        style_header(ws_constraints, constraint_headers)

        ws_constraints.append(["max_consecutive_hours", 3, "最大连排课时"])
        ws_constraints.append(["main_subject_prefer_morning", "是", "主课优先上午"])
        ws_constraints.append(["min_course_interval_days", 1, "同课程最小间隔天数"])
        ws_constraints.append(["max_daily_hours_per_teacher", 4, "教师每天最大课时"])
        ws_constraints.append(["course_even_distribution", "是", "课程均匀分布"])

        for col_letter in ['A', 'B', 'C']:
            ws_constraints.column_dimensions[col_letter].width = 25

        # 保存文件并关闭工作簿
        wb.save(output_path)
        wb.close()
        logger.info(f"模板已生成: {output_path}")
        return True

    except Exception as e:
        logger.error(f"生成模板失败: {e}")
        return False


def parse_scheduling_excel(file_path: str) -> Optional[Dict]:
    """
    解析排课数据 Excel 文件

    参数:
        file_path: Excel 文件路径

    返回:
        解析后的数据字典，失败返回 None
    """
    if not HAS_OPENPYXL:
        logger.error("openpyxl 未安装")
        return None

    try:
        wb = openpyxl.load_workbook(file_path)
        data = {
            "classes": [],
            "teachers": [],
            "courses": [],
            "classrooms": [],
            "constraints": {},
        }

        # 解析班级信息
        if "班级信息" in wb.sheetnames:
            ws = wb["班级信息"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # 有数据
                    # 解析固定教室（第7列，索引6）
                    assigned_classrooms = []
                    if len(row) > 6 and row[6]:
                        assigned_classrooms = [c.strip() for c in str(row[6]).split(",") if c.strip()]

                    class_data = {
                        "id": str(row[0]),
                        "name": str(row[1]) if row[1] else "",
                        "grade": str(row[2]) if row[2] else "",
                        "student_count": int(row[3]) if row[3] else 45,
                        "homeroom_teacher": str(row[4]) if row[4] else "",
                        "courses": [c.strip() for c in str(row[5]).split(",")] if row[5] else [],
                        "assigned_classrooms": assigned_classrooms,
                    }
                    data["classes"].append(class_data)

        # 解析教师信息
        if "教师信息" in wb.sheetnames:
            ws = wb["教师信息"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    teacher_data = {
                        "id": str(row[0]),
                        "name": str(row[1]) if row[1] else "",
                        "subjects": [s.strip() for s in str(row[2]).split(",")] if row[2] else [],
                        "max_hours_per_day": int(row[3]) if row[3] else 4,
                        "max_hours_per_week": int(row[4]) if row[4] else 20,
                    }
                    data["teachers"].append(teacher_data)

        # 解析课程信息
        if "课程信息" in wb.sheetnames:
            ws = wb["课程信息"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    course_data = {
                        "id": str(row[0]),
                        "name": str(row[1]) if row[1] else "",
                        "subject": str(row[2]) if row[2] else "",
                        "hours_per_week": int(row[3]) if row[3] else 2,
                        "is_main_subject": str(row[4]).lower() in ["是", "yes", "true"] if row[4] else False,
                        "needs_consecutive": str(row[5]).lower() in ["是", "yes", "true"] if row[5] else False,
                        "required_equipment": [e.strip() for e in str(row[6]).split(",")] if row[6] else [],
                    }
                    data["courses"].append(course_data)

        # 解析教室信息
        if "教室信息" in wb.sheetnames:
            ws = wb["教室信息"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    classroom_data = {
                        "id": str(row[0]),
                        "name": str(row[1]) if row[1] else "",
                        "capacity": int(row[2]) if row[2] else 50,
                        "equipment": [e.strip() for e in str(row[3]).split(",")] if row[3] else [],
                        "building": str(row[4]) if row[4] else "",
                    }
                    data["classrooms"].append(classroom_data)

        # 解析约束配置（支持两种工作表名称）
        constraint_sheet = None
        if "约束配置" in wb.sheetnames:
            constraint_sheet = "约束配置"
        elif "约束条件" in wb.sheetnames:
            constraint_sheet = "约束条件"

        if constraint_sheet:
            ws = wb[constraint_sheet]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1] is not None:
                    key = str(row[0])
                    value = row[1]
                    # 转换布尔值
                    if isinstance(value, str):
                        if value.lower() in ["是", "yes", "true"]:
                            value = True
                        elif value.lower() in ["否", "no", "false"]:
                            value = False
                    data["constraints"][key] = value

        logger.info(f"Excel 解析完成: {len(data['classes'])} 班级, "
                    f"{len(data['teachers'])} 教师, {len(data['courses'])} 课程")
        return data

    except Exception as e:
        logger.error(f"解析 Excel 失败: {e}")
        return None


def export_schedule_to_excel(schedule, classes: Dict, courses: Dict,
                             teachers: Dict, output_path: str,
                             classrooms: Dict = None) -> bool:
    """
    导出课表到 Excel

    参数:
        schedule: Schedule 对象
        classes: 班级字典
        courses: 课程字典
        teachers: 教师字典
        output_path: 输出路径

    返回:
        是否成功
    """
    if not HAS_OPENPYXL:
        logger.error("openpyxl 未安装")
        return False

    try:
        wb = openpyxl.Workbook()

        # 样式定义
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        weekdays = ["周一", "周二", "周三", "周四", "周五"]
        periods = [("第1节", 1), ("第2节", 2), ("第3节", 3), ("第4节", 4),
                   ("第5节", 5), ("第6节", 6), ("第7节", 7), ("第8节", 8)]

        # 为每个班级生成一个工作表
        for class_id, class_group in classes.items():
            # 工作表名称（限制31字符）
            sheet_name = class_group.name[:31]
            ws = wb.create_sheet(sheet_name)

            # 标题行
            ws.merge_cells('A1:F1')
            title_cell = ws.cell(row=1, column=1, value=f"{class_group.name} 课表")
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal="center")

            # 表头
            headers = ["节次"] + weekdays
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col, value=header)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # 获取班级课表
            class_schedule = schedule.get_class_schedule(class_id)

            # 填充课表（按单节显示）
            for row_idx, (period_name, period_num) in enumerate(periods, 3):
                # 节次列
                cell = ws.cell(row=row_idx, column=1, value=period_name)
                cell.font = header_font
                cell.alignment = cell_alignment
                cell.border = thin_border

                # 每天的课程
                for col_idx, weekday in enumerate(weekdays, 2):
                    # 查找该节次的课程
                    course_text = ""
                    for entry in class_schedule:
                        if (entry.time_slot.weekday.value == weekday and
                                entry.time_slot.period == period_num):
                            course = courses.get(entry.course_id)
                            teacher = teachers.get(entry.teacher_id)
                            classroom = classrooms.get(entry.classroom_id) if classrooms else None
                            if course and teacher:
                                course_text = f"{course.name}\n({teacher.name})"
                                if classroom:
                                    course_text += f"\n{classroom.name}"
                            break

                    cell = ws.cell(row=row_idx, column=col_idx, value=course_text)
                    cell.alignment = cell_alignment
                    cell.border = thin_border

            # 设置列宽
            ws.column_dimensions['A'].width = 10
            for col_letter in ['B', 'C', 'D', 'E', 'F']:
                ws.column_dimensions[col_letter].width = 20

            # 设置行高
            for row in range(3, 11):
                ws.row_dimensions[row].height = 50

        # 删除默认工作表
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # 保存文件并关闭工作簿（释放文件句柄）
        wb.save(output_path)
        wb.close()
        logger.info(f"课表已导出: {output_path}")
        return True

    except Exception as e:
        logger.error(f"导出课表失败: {e}")
        return False
