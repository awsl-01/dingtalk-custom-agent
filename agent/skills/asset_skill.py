"""
资产管理技能 - 支持资产的录入、查询、借还、统计、模板导入

使用示例：
- "录入资产 投影仪 教学设备 301教室"
- "查询资产 投影仪"
- "借用资产 AST20260618001 李老师"
- "归还资产 AST20260618001"
- "资产统计"
- "资产录入模板" - 下载资产批量导入模板
- "资产盘点模板" - 下载资产盘点模板
"""
import os
import re
import csv
from typing import Dict, List
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from .registry import BaseSkill, skill_registry
from .asset_storage import (
    load_assets, create_asset, get_asset_by_id, get_assets_by_name,
    get_asset_stats, add_borrow_record, return_asset, search_assets,
    save_assets
)

# 模板输出目录
TEMPLATES_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "templates", "asset")


class AssetSkill(BaseSkill):
    """资产管理技能"""

    @property
    def name(self) -> str:
        return "资产管理"

    @property
    def description(self) -> str:
        return "管理学校资产：录入、查询、借还、统计"

    @property
    def keywords(self) -> list:
        return [
            "资产", "设备", "录入资产", "查询资产", "借用", "归还",
            "资产统计", "资产列表", "设备管理", "设备查询",
            "投影仪", "电脑", "桌椅", "教材", "器材",
            "资产模板", "资产录入模板", "资产盘点模板", "批量导入"
        ]

    @property
    def priority(self) -> int:
        return 50

    def can_handle(self, text: str) -> float:
        """判断是否能处理资产管理"""
        text_lower = text.lower()

        # 模板相关关键词（高优先级）
        template_keywords = ["资产模板", "资产录入模板", "资产盘点模板", "批量导入模板"]
        for kw in template_keywords:
            if kw in text_lower:
                return 0.95

        # 精确匹配关键词
        exact_keywords = ["录入资产", "添加资产", "新增资产"]
        for kw in exact_keywords:
            if kw in text_lower:
                return 0.9

        query_keywords = ["查询资产", "资产查询", "查找资产", "资产列表", "设备列表"]
        for kw in query_keywords:
            if kw in text_lower:
                return 0.9

        borrow_keywords = ["借用资产", "借出资产", "设备借用"]
        for kw in borrow_keywords:
            if kw in text_lower:
                return 0.9

        return_keywords = ["归还资产", "设备归还", "还回资产"]
        for kw in return_keywords:
            if kw in text_lower:
                return 0.9

        stats_keywords = ["资产统计", "资产报表", "设备统计", "资产概况"]
        for kw in stats_keywords:
            if kw in text_lower:
                return 0.9

        # 导入相关关键词
        import_keywords = ["导入资产", "批量导入", "批量录入"]
        for kw in import_keywords:
            if kw in text_lower:
                return 0.85

        # 通用关键词匹配
        general_keywords = ["资产", "设备"]
        for kw in general_keywords:
            if kw in text_lower:
                return 0.7

        return 0

    def extract_info(self, text: str) -> dict:
        """提取资产相关信息"""
        info = {
            "action": None,
            "name": None,
            "category": None,
            "location": None,
            "asset_id": None,
            "borrower": None,
            "status": None,
            "template_type": None
        }

        # 识别操作类型
        # 模板相关（高优先级）
        if any(kw in text for kw in ["资产录入模板", "批量导入模板", "资产模板"]):
            info["action"] = "template"
            info["template_type"] = "import"
        elif any(kw in text for kw in ["资产盘点模板", "盘点模板"]):
            info["action"] = "template"
            info["template_type"] = "inventory"
        elif any(kw in text for kw in ["导入资产", "批量导入", "批量录入"]):
            info["action"] = "batch_import"
        elif any(kw in text for kw in ["录入", "添加", "新增"]):
            info["action"] = "add"
        elif any(kw in text for kw in ["查询", "查找", "列表"]):
            info["action"] = "query"
        elif any(kw in text for kw in ["借用", "借出"]):
            info["action"] = "borrow"
        elif any(kw in text for kw in ["归还", "还回"]):
            info["action"] = "return"
        elif any(kw in text for kw in ["统计", "报表", "概况"]):
            info["action"] = "stats"

        # 提取资产名称（常见设备）
        device_keywords = [
            "投影仪", "电脑", "笔记本", "平板", "桌椅", "课桌", "椅子",
            "教材", "器材", "实验设备", "音响", "麦克风", "摄像头",
            "打印机", "复印机", "扫描仪", "空调", "电视", "白板"
        ]
        for device in device_keywords:
            if device in text:
                info["name"] = device
                break

        # 提取资产编号（AST开头）
        id_match = re.search(r'(AST\d{11})', text, re.IGNORECASE)
        if id_match:
            info["asset_id"] = id_match.group(1).upper()

        # 提取教室位置
        location_match = re.search(r'(\d{3})教室', text)
        if location_match:
            info["location"] = location_match.group(1) + "教室"

        # 提取分类
        category_keywords = {
            "教学设备": ["投影仪", "电脑", "白板", "音响", "麦克风"],
            "办公设备": ["打印机", "复印机", "扫描仪", "电脑"],
            "家具": ["桌椅", "课桌", "椅子", "柜子"],
            "教材": ["教材", "课本", "教辅"]
        }
        for category, keywords in category_keywords.items():
            for kw in keywords:
                if kw in text:
                    info["category"] = category
                    break
            if info["category"]:
                break

        # 提取借用人（X老师）
        borrower_match = re.search(r'([^\s]{2,4})老师', text)
        if borrower_match:
            info["borrower"] = borrower_match.group(1) + "老师"

        return info

    async def execute(self, text: str, context: dict) -> str:
        """执行资产管理操作"""
        info = self.extract_info(text)
        corp_id = context.get("corp_id", "default")
        action = info.get("action")

        if action == "template":
            return self._handle_template(info, context)
        elif action == "batch_import":
            return await self._batch_import(info, corp_id, context)
        elif action == "add":
            return await self._add_asset(info, corp_id, context)
        elif action == "query":
            return await self._query_asset(info, corp_id)
        elif action == "borrow":
            return await self._borrow_asset(info, corp_id)
        elif action == "return":
            return await self._return_asset(info, corp_id)
        elif action == "stats":
            return await self._get_stats(corp_id)
        else:
            return self._get_help()

    async def _add_asset(self, info: dict, corp_id: str, context: dict = None) -> str:
        """添加新资产"""
        name = info.get("name")
        category = info.get("category", "其他")
        location = info.get("location", "未指定")

        if not name:
            return "请提供资产名称，例如：录入资产 投影仪 教学设备 301教室"

        # 使用存储模块创建资产
        asset_data = {
            "name": name,
            "category": category,
            "location": location,
            "responsible_user": (context or {}).get("sender_nick", "未指定"),
            "description": f"{name} - {category}"
        }

        new_asset = create_asset(asset_data, corp_id)

        return f"""[成功] 资产录入成功！

[资产信息]
- 编号：{new_asset['id']}
- 名称：{new_asset['name']}
- 分类：{new_asset['category']}
- 位置：{new_asset['location']}
- 状态：在用
- 录入时间：{new_asset['created_at']}

[提示] 可使用「查询资产 {name}」查看资产详情"""

    async def _query_asset(self, info: dict, corp_id: str) -> str:
        """查询资产"""
        name = info.get("name")
        asset_id = info.get("asset_id")

        # 按编号查询
        if asset_id:
            asset = get_asset_by_id(asset_id, corp_id)
            if asset:
                return self._format_asset_detail(asset)
            return f"[错误] 未找到编号为 {asset_id} 的资产"

        # 按名称查询
        if name:
            found = get_assets_by_name(name, corp_id)
            if not found:
                return f"[错误] 未找到包含「{name}」的资产"

            result = f"[查询结果] 查询到 {len(found)} 条资产记录：\n\n"
            for asset in found[:10]:  # 最多显示10条
                result += f"- {asset['id']} | {asset['name']} | {asset['status']} | {asset['location']}\n"
            if len(found) > 10:
                result += f"\n... 还有 {len(found) - 10} 条记录"
            return result

        # 显示所有资产（最多20条）
        assets = load_assets(corp_id)
        if not assets:
            return "[提示] 暂无资产记录，请先使用「录入资产」添加资产"

        result = f"[资产列表] 共有 {len(assets)} 条资产记录：\n\n"
        for asset in assets[:20]:
            result += f"- {asset['id']} | {asset['name']} | {asset['status']} | {asset['location']}\n"
        if len(assets) > 20:
            result += f"\n... 还有 {len(assets) - 20} 条记录"
        return result

    async def _borrow_asset(self, info: dict, corp_id: str) -> str:
        """借用资产"""
        asset_id = info.get("asset_id")
        borrower = info.get("borrower")

        if not asset_id:
            return "请提供资产编号，例如：借用资产 AST20260618001 李老师"
        if not borrower:
            return "请提供借用人，例如：借用资产 AST20260618001 李老师"

        # 检查资产是否存在
        asset = get_asset_by_id(asset_id, corp_id)
        if not asset:
            return f"[错误] 未找到编号为 {asset_id} 的资产"

        if asset.get('status') != '在用':
            return f"[错误] 资产 {asset_id} 当前状态为「{asset.get('status')}」，无法借用"

        # 添加借用记录
        borrow_record = add_borrow_record(asset_id, borrower, corp_id)

        if borrow_record:
            return f"""[成功] 资产借用成功！

[借用信息]
- 资产编号：{asset_id}
- 资产名称：{asset.get('name')}
- 借用人：{borrower}
- 借用时间：{borrow_record['borrow_date']}

[提示] 请记得及时归还！"""

        return f"[错误] 借用失败，请稍后重试"

    async def _return_asset(self, info: dict, corp_id: str) -> str:
        """归还资产"""
        asset_id = info.get("asset_id")

        if not asset_id:
            return "请提供资产编号，例如：归还资产 AST20260618001"

        # 检查资产是否存在
        asset = get_asset_by_id(asset_id, corp_id)
        if not asset:
            return f"[错误] 未找到编号为 {asset_id} 的资产"

        if asset.get('status') != '借用中':
            return f"[错误] 资产 {asset_id} 当前状态为「{asset.get('status')}」，无需归还"

        # 执行归还
        success = return_asset(asset_id, corp_id)

        if success:
            # 重新获取资产信息以获取更新时间
            updated_asset = get_asset_by_id(asset_id, corp_id)
            return f"""[成功] 资产归还成功！

[归还信息]
- 资产编号：{asset_id}
- 资产名称：{asset.get('name')}
- 归还时间：{updated_asset.get('updated_at')}

感谢使用！"""

        return f"[错误] 归还失败，请稍后重试"

    async def _get_stats(self, corp_id: str) -> str:
        """获取资产统计"""
        stats = get_asset_stats(corp_id)

        if stats["total"] == 0:
            return "[提示] 暂无资产记录"

        result = f"""[资产统计报告]

[总资产数] {stats['total']} 台

[状态分布]
"""
        for status, count in sorted(stats["status_count"].items()):
            result += f"- {status}：{count} 台\n"

        result += "\n[分类分布]\n"
        for category, count in sorted(stats["category_count"].items(), key=lambda x: -x[1]):
            result += f"- {category}：{count} 台\n"

        result += "\n[位置分布（前5）]\n"
        for location, count in sorted(stats["location_count"].items(), key=lambda x: -x[1])[:5]:
            result += f"- {location}：{count} 台\n"

        return result

    def _format_asset_detail(self, asset: Dict) -> str:
        """格式化资产详情"""
        result = f"""[资产详情]

- 编号：{asset.get('id')}
- 名称：{asset.get('name')}
- 分类：{asset.get('category')}
- 位置：{asset.get('location')}
- 状态：{asset.get('status')}
- 负责人：{asset.get('responsible_user')}
- 采购日期：{asset.get('purchase_date')}
- 创建时间：{asset.get('created_at')}
- 更新时间：{asset.get('updated_at')}"""

        if asset.get('borrow_records'):
            last_record = asset['borrow_records'][-1]
            if last_record.get('return_date') is None:
                result += f"""

[当前借用]
- 借用人：{last_record.get('borrower')}
- 借用时间：{last_record.get('borrow_date')}"""

        return result

    def _get_help(self) -> str:
        """获取帮助信息"""
        return """[资产管理使用指南]

[1. 录入资产]
  格式：录入资产 [名称] [分类] [位置]
  示例：录入资产 投影仪 教学设备 301教室

[2. 查询资产]
  格式：查询资产 [名称/编号]
  示例：查询资产 投影仪
        查询资产 AST20260618001

[3. 借用资产]
  格式：借用资产 [编号] [借用人]
  示例：借用资产 AST20260618001 李老师

[4. 归还资产]
  格式：归还资产 [编号]
  示例：归还资产 AST20260618001

[5. 资产统计]
  格式：资产统计

[6. 模板下载]
  格式：资产录入模板 / 资产盘点模板
  说明：下载 Excel 模板，填写后可批量导入

[7. 批量导入]
  格式：批量导入资产
  说明：上传填写好的模板文件进行批量导入

[提示] 资产编号以 AST 开头，共14位"""

    def _handle_template(self, info: dict, context: dict) -> str:
        """处理模板下载请求"""
        template_type = info.get("template_type", "import")

        # 确保模板目录存在
        os.makedirs(TEMPLATES_DIR, exist_ok=True)

        if template_type == "import":
            return self._create_import_template(context)
        elif template_type == "inventory":
            return self._create_inventory_template(context)
        else:
            return "[错误] 未知的模板类型"

    def _create_import_template(self, context: dict) -> str:
        """创建资产录入模板"""
        if HAS_OPENPYXL:
            filepath = os.path.join(TEMPLATES_DIR, "资产录入模板.xlsx")
            self._create_import_excel_template(filepath)
            filename = "资产录入模板.xlsx"
        else:
            filepath = os.path.join(TEMPLATES_DIR, "资产录入模板.csv")
            self._create_import_csv_template(filepath)
            filename = "资产录入模板.csv"

        context["_file_to_send"] = filepath
        context["_file_name"] = filename
        context["_file_type"] = "file"

        return (
            "[资产录入模板已生成]\n\n"
            "[模板字段说明]\n"
            "- 资产名称：必填\n"
            "- 资产分类：教学设备/办公设备/家具/教材/其他\n"
            "- 存放位置：必填（如：301教室、办公室）\n"
            "- 负责人：选填\n"
            "- 采购日期：选填（格式：2026-06-18）\n"
            "- 资产描述：选填\n\n"
            "[使用说明]\n"
            "1. 下载并打开模板文件\n"
            "2. 按照字段说明填写资产信息\n"
            "3. 保存文件后上传\n"
            "4. 发送「批量导入资产」进行导入\n\n"
            "[示例数据]\n"
            "- 投影仪 | 教学设备 | 301教室 | 张老师 | 2025-09-01 | 索尼投影仪\n"
            "- 电脑 | 办公设备 | 办公室 | 李老师 | 2025-08-15 | 联想台式机"
        )

    def _create_import_excel_template(self, filepath: str):
        """创建 Excel 格式的资产录入模板"""
        wb = Workbook()
        ws = wb.active
        ws.title = "资产录入"

        # 定义样式
        header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 设置表头
        headers = ['资产名称', '资产分类', '存放位置', '负责人', '采购日期', '资产描述']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 设置列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 30

        # 添加示例数据
        examples = [
            ['投影仪', '教学设备', '301教室', '张老师', '2025-09-01', '索尼投影仪'],
            ['电脑', '办公设备', '办公室', '李老师', '2025-08-15', '联想台式机'],
            ['课桌', '家具', '201教室', '', '2025-07-01', '学生课桌'],
        ]
        for row_idx, example in enumerate(examples, 2):
            for col_idx, value in enumerate(example, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # 添加数据验证说明
        ws_note = wb.create_sheet("填写说明")
        notes = [
            ['字段名', '是否必填', '说明', '示例'],
            ['资产名称', '是', '资产的名称', '投影仪'],
            ['资产分类', '是', '教学设备/办公设备/家具/教材/其他', '教学设备'],
            ['存放位置', '是', '资产存放的位置', '301教室'],
            ['负责人', '否', '资产负责人', '张老师'],
            ['采购日期', '否', '采购日期，格式：YYYY-MM-DD', '2025-09-01'],
            ['资产描述', '否', '资产的补充描述', '索尼投影仪'],
        ]
        for row_idx, note in enumerate(notes, 1):
            for col_idx, value in enumerate(note, 1):
                ws_note.cell(row=row_idx, column=col_idx, value=value)

        # 设置说明表的列宽
        ws_note.column_dimensions['A'].width = 15
        ws_note.column_dimensions['B'].width = 12
        ws_note.column_dimensions['C'].width = 35
        ws_note.column_dimensions['D'].width = 20

        wb.save(filepath)

    def _create_import_csv_template(self, filepath: str):
        """创建 CSV 格式的资产录入模板（备用）"""
        with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["资产名称", "资产分类", "存放位置", "负责人", "采购日期", "资产描述"])
            writer.writerow(["投影仪", "教学设备", "301教室", "张老师", "2025-09-01", "索尼投影仪"])
            writer.writerow(["电脑", "办公设备", "办公室", "李老师", "2025-08-15", "联想台式机"])
            writer.writerow(["课桌", "家具", "201教室", "", "2025-07-01", "学生课桌"])

    def _create_inventory_template(self, context: dict) -> str:
        """创建资产盘点模板"""
        if HAS_OPENPYXL:
            filepath = os.path.join(TEMPLATES_DIR, "资产盘点模板.xlsx")
            self._create_inventory_excel_template(filepath)
            filename = "资产盘点模板.xlsx"
        else:
            filepath = os.path.join(TEMPLATES_DIR, "资产盘点模板.csv")
            self._create_inventory_csv_template(filepath)
            filename = "资产盘点模板.csv"

        context["_file_to_send"] = filepath
        context["_file_name"] = filename
        context["_file_type"] = "file"

        return (
            "[资产盘点模板已生成]\n\n"
            "[模板字段说明]\n"
            "- 资产编号：必填（AST开头的编号）\n"
            "- 资产名称：必填\n"
            "- 实际位置：必填（盘点时的实际位置）\n"
            "- 资产状态：在用/闲置/维修中/报废\n"
            "- 盘点人：必填\n"
            "- 盘点日期：必填（格式：2026-06-18）\n"
            "- 备注：选填\n\n"
            "[使用说明]\n"
            "1. 先使用「资产列表」导出当前资产清单\n"
            "2. 下载盘点模板并填写实际盘点结果\n"
            "3. 上传填写好的模板文件\n"
            "4. 系统将自动更新资产状态\n\n"
            "[注意事项]\n"
            "- 请确保资产编号正确\n"
            "- 如发现资产丢失或损坏，请在备注中说明"
        )

    def _create_inventory_excel_template(self, filepath: str):
        """创建 Excel 格式的资产盘点模板"""
        wb = Workbook()
        ws = wb.active
        ws.title = "资产盘点"

        # 定义样式
        header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 设置表头
        headers = ['资产编号', '资产名称', '实际位置', '资产状态', '盘点人', '盘点日期', '备注']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 设置列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 30

        # 添加示例数据
        examples = [
            ['AST20260618001', '投影仪', '301教室', '在用', '王老师', '2026-06-18', ''],
            ['AST20260618002', '电脑', '办公室', '在用', '王老师', '2026-06-18', ''],
            ['AST20260618003', '课桌', '仓库', '闲置', '王老师', '2026-06-18', '已搬迁'],
        ]
        for row_idx, example in enumerate(examples, 2):
            for col_idx, value in enumerate(example, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # 添加数据验证说明
        ws_note = wb.create_sheet("填写说明")
        notes = [
            ['字段名', '是否必填', '说明', '示例'],
            ['资产编号', '是', 'AST开头的资产编号', 'AST20260618001'],
            ['资产名称', '是', '资产的名称', '投影仪'],
            ['实际位置', '是', '盘点时资产的实际位置', '301教室'],
            ['资产状态', '是', '在用/闲置/维修中/报废', '在用'],
            ['盘点人', '是', '执行盘点的人员', '王老师'],
            ['盘点日期', '是', '盘点日期，格式：YYYY-MM-DD', '2026-06-18'],
            ['备注', '否', '补充说明', '已搬迁、损坏等'],
        ]
        for row_idx, note in enumerate(notes, 1):
            for col_idx, value in enumerate(note, 1):
                ws_note.cell(row=row_idx, column=col_idx, value=value)

        # 设置说明表的列宽
        ws_note.column_dimensions['A'].width = 15
        ws_note.column_dimensions['B'].width = 12
        ws_note.column_dimensions['C'].width = 35
        ws_note.column_dimensions['D'].width = 25

        wb.save(filepath)

    def _create_inventory_csv_template(self, filepath: str):
        """创建 CSV 格式的资产盘点模板（备用）"""
        with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["资产编号", "资产名称", "实际位置", "资产状态", "盘点人", "盘点日期", "备注"])
            writer.writerow(["AST20260618001", "投影仪", "301教室", "在用", "王老师", "2026-06-18", ""])
            writer.writerow(["AST20260618002", "电脑", "办公室", "在用", "王老师", "2026-06-18", ""])
            writer.writerow(["AST20260618003", "课桌", "仓库", "闲置", "王老师", "2026-06-18", "已搬迁"])

    async def _batch_import(self, info: dict, corp_id: str, context: dict) -> str:
        """批量导入资产"""
        # 检查是否有文件上传
        file_path = context.get("_uploaded_file")
        if not file_path:
            return (
                "[提示] 请先上传填写好的资产模板文件\n\n"
                "[操作步骤]\n"
                "1. 发送「资产录入模板」下载模板\n"
                "2. 按照模板填写资产信息\n"
                "3. 上传填写好的文件\n"
                "4. 发送「批量导入资产」完成导入"
            )

        # 检查文件类型
        if not file_path.endswith(('.xlsx', '.csv')):
            return "[错误] 仅支持 .xlsx 或 .csv 格式的文件"

        try:
            if file_path.endswith('.xlsx'):
                imported_count = self._import_from_excel(file_path, corp_id)
            else:
                imported_count = self._import_from_csv(file_path, corp_id)

            return (
                f"[批量导入完成]\n\n"
                f"[导入结果]\n"
                f"- 成功导入：{imported_count} 条资产记录\n"
                f"- 数据已保存到知识库\n\n"
                f"[提示] 使用「资产列表」查看导入的资产"
            )
        except Exception as e:
            return f"[错误] 导入失败：{str(e)}"

    def _import_from_excel(self, file_path: str, corp_id: str) -> int:
        """从 Excel 文件导入资产"""
        from openpyxl import load_workbook

        wb = load_workbook(file_path)
        ws = wb.active

        imported_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # 跳过空行
                continue

            asset_data = {
                "name": str(row[0] or ""),
                "category": str(row[1] or "其他"),
                "location": str(row[2] or "未指定"),
                "responsible_user": str(row[3] or "未指定"),
                "purchase_date": str(row[4] or datetime.now().strftime("%Y-%m-%d")),
                "description": str(row[5] or "")
            }

            if asset_data["name"]:  # 确保有资产名称
                create_asset(asset_data, corp_id)
                imported_count += 1

        return imported_count

    def _import_from_csv(self, file_path: str, corp_id: str) -> int:
        """从 CSV 文件导入资产"""
        imported_count = 0

        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            next(reader)  # 跳过表头

            for row in reader:
                if not row or not row[0]:  # 跳过空行
                    continue

                asset_data = {
                    "name": str(row[0] or ""),
                    "category": str(row[1] or "其他"),
                    "location": str(row[2] or "未指定"),
                    "responsible_user": str(row[3] or "未指定"),
                    "purchase_date": str(row[4] or datetime.now().strftime("%Y-%m-%d")),
                    "description": str(row[5] or "")
                }

                if asset_data["name"]:  # 确保有资产名称
                    create_asset(asset_data, corp_id)
                    imported_count += 1

        return imported_count


# 注册技能
skill_registry.register(AssetSkill())
