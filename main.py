import logging
import sys
import os
import json
import tempfile
import requests
import re
import asyncio
import concurrent.futures
import hashlib
import time
from datetime import datetime
from typing import Tuple

import dingtalk_stream
from dingtalk_stream import AckMessage
from dingtalk_stream.chatbot import ChatbotHandler, ChatbotMessage

import config
from agent.core import chat, chat_with_knowledge
from agent.ppt_master_integration import generate_ppt_with_master
from agent.web_search import (
    search_web,
    search_and_summarize,
    search_for_education,
    quick_search,
    search_teaching_resources,
    search_exam_questions,
    search_materials
)
from agent.conversation_state import state_manager, TaskType, TaskStatus
from agent.school_config import school_manager, SchoolConfig
from agent.knowledge_base_v2 import get_knowledge_base
from agent.media_handler import (
    download_media_from_dingtalk,
    extract_text_from_file,
    get_today_dir,
)
from agent.structured_data import (
    ScheduleParser, ExamParser, ContactParser,
    detect_data_type,
)
from agent.skills import skill_registry
from agent.skills.loader import load_skills
from agent.intent_router import intent_router, UserIntent

# Web 日志记录（延迟导入避免循环依赖）
_web_log_enabled = True
try:
    from web.models import SessionLocal, MessageLog
except ImportError:
    _web_log_enabled = False
    logger = logging.getLogger(__name__)
    logger.warning("Web 日志模块未找到，消息日志将不记录到数据库")

# 强制使用 UTF-8 编码
sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

logging.basicConfig(level=logging.DEBUG, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

# 加载技能模块
loaded_skills = load_skills()
logger.info(f"已加载技能: {loaded_skills}")

# PPT生成关键词
PPT_KEYWORDS = ["ppt", "PPT", "幻灯片", "演示文稿", "slides"]

# 教育PPT关键词
EDUCATION_KEYWORDS = ["教案", "课件", "说课", "反思", "教学设计", "教学大纲", "学情", "难度"]

# 年级关键词
GRADE_KEYWORDS = ["小学", "初中", "高中", "大学", "年级", "学前", "幼儿园",
                  "初一", "初二", "初三", "高一", "高二", "高三",
                  "一年级", "二年级", "三年级", "四年级", "五年级", "六年级"]

# 学科关键词
SUBJECT_KEYWORDS = ["语文", "数学", "英语", "物理", "化学", "生物", "历史", "地理", "政治",
                    "音乐", "美术", "体育", "科学", "信息技术", "心理", "班会", "德育"]

# 搜索相关关键词
SEARCH_KEYWORDS = ["搜索", "查询", "查找", "找一下", "搜一下", "帮我查", "帮我找", "百度", "谷歌"]
RESOURCE_KEYWORDS = ["资源", "素材", "模板", "习题", "试题", "练习题", "视频", "动画"]
NEWS_KEYWORDS = ["新闻", "最新", "最近", "今日", "今天", "热点"]


def is_ppt_request(text: str) -> bool:
    """判断是否是PPT生成请求"""
    text_lower = text.lower()
    return any(keyword in text_lower for keyword in PPT_KEYWORDS)


def is_education_request(text: str) -> bool:
    """判断是否是教育相关请求"""
    text_lower = text.lower()
    if any(keyword in text_lower for keyword in EDUCATION_KEYWORDS):
        return True
    if any(keyword in text_lower for keyword in GRADE_KEYWORDS):
        return True
    if any(keyword in text_lower for keyword in SUBJECT_KEYWORDS):
        return True
    return False


def is_search_request(text: str) -> bool:
    text_lower = text.lower()
    return any(keyword in text_lower for keyword in SEARCH_KEYWORDS)


def is_resource_request(text: str) -> bool:
    text_lower = text.lower()
    return any(keyword in text_lower for keyword in RESOURCE_KEYWORDS)


def is_news_request(text: str) -> bool:
    text_lower = text.lower()
    return any(keyword in text_lower for keyword in NEWS_KEYWORDS)


def extract_search_query(text: str) -> str:
    query = text
    for keyword in SEARCH_KEYWORDS + ["请", "帮我", "一下", "关于", "的"]:
        query = query.replace(keyword, "")
    query = ' '.join(query.split())
    query = query.strip() if query.strip() else text
    return query


def parse_education_info(text: str) -> dict:
    """解析教育PPT请求中的信息"""
    info = {'topic': '', 'subject': '', 'grade': '', 'difficulty': '中等', 'content_type': '课件', 'chapter': '', 'page_count': None}

    subjects = ['语文', '数学', '英语', '物理', '化学', '生物', '历史', '地理', '政治', '音乐', '美术', '体育']
    for subject in subjects:
        if subject in text:
            info['subject'] = subject
            break

    grade_patterns = [
        r'小学[一二三四五六]年级', r'初中[一二三]', r'高中[一二三]',
        r'[一二三四五六七八九十]年级', r'高[一二三]', r'初[一二三]',
    ]
    for pattern in grade_patterns:
        match = re.search(pattern, text)
        if match:
            info['grade'] = match.group()
            break

    difficulties = ['基础', '中等', '提高', '拓展']
    for diff in difficulties:
        if diff in text:
            info['difficulty'] = diff
            break

    if '教案' in text:
        info['content_type'] = '教案'
    elif '说课' in text:
        info['content_type'] = '说课稿'
    elif '反思' in text:
        info['content_type'] = '反思'
    elif '课件' in text or '大纲' in text:
        info['content_type'] = '课件'

    # 提取页数（如"15页PPT"、"20页"、"15页的"）
    page_count_match = re.search(r'(\d+)\s*页', text)
    if page_count_match:
        info['page_count'] = int(page_count_match.group(1))

    # 提取主题和章节
    topic_match = re.search(r'[《](.*?)[》]', text)
    if topic_match:
        info['topic'] = topic_match.group(1)
        info['chapter'] = topic_match.group(1)
    else:
        clean_text = text
        for keyword in ['生成', '制作', '请', '帮我', '教案', '课件', '说课', '反思',
                        '教学设计', '难度', '基础', '中等', '提高', '拓展',
                        '一个', '适合', '的', 'PPT', 'ppt', '幻灯片']:
            clean_text = re.sub(keyword, '', clean_text)
        for subject in subjects:
            clean_text = clean_text.replace(subject, '')
        for pattern in grade_patterns:
            clean_text = re.sub(pattern, '', clean_text)
        # 移除页数描述
        clean_text = re.sub(r'\d+\s*页', '', clean_text)
        info['topic'] = clean_text.strip()[:30] if clean_text.strip() else '教学内容'
        info['chapter'] = info['topic']

    return info


def generate_ppt_outline_markdown(topic: str, subject: str = "", grade: str = "",
                                  search_context: str = "") -> str:
    """生成PPT大纲供用户确认"""
    client = __import__('openai').OpenAI(
        api_key=config.OPENAI_API_KEY,
        base_url=config.OPENAI_BASE_URL,
    )

    system_prompt = """你是一个专业的教学内容设计师。请根据用户提供的主题，生成一个详细的PPT大纲（Markdown格式）。

大纲要求：
1. 标题清晰明确
2. 包含10-15个主要章节
3. 每个章节下有3-5个要点
4. 内容准确、结构合理
5. 适合对应学段的教学需求
6. 如有教材参考内容，请优先基于真实知识点组织大纲

请直接输出Markdown格式的大纲，不要添加额外说明。"""

    user_message = f"""请为以下主题生成PPT大纲：

主题：{topic}
学科：{subject if subject else '未指定'}
年级：{grade if grade else '未指定'}

{f'教材参考内容（基于网络搜索）：\n{search_context}\n请基于以上内容组织大纲，确保知识点准确。' if search_context else ''}

请生成详细的Markdown大纲。"""

    try:
        response = client.chat.completions.create(
            model=config.OPENAI_MODEL,
            max_tokens=2048,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_message},
            ],
        )
        return response.choices[0].message.content
    except Exception as e:
        logger.error(f"生成大纲失败: {e}")
        return f"# {topic}\n\n## 一、导入\n\n## 二、主要内容\n\n## 三、总结"


def get_access_token() -> str:
    token_url = "https://oapi.dingtalk.com/gettoken"
    params = {
        "appkey": config.DINGTALK_APP_KEY,
        "appsecret": config.DINGTALK_APP_SECRET,
    }
    resp = requests.get(token_url, params=params)
    token_data = resp.json()
    if token_data.get("errcode") != 0:
        raise RuntimeError(f"获取token失败: {token_data}")
    return token_data["access_token"]


def upload_media(access_token: str, file_path: str) -> str:
    upload_url = f"https://oapi.dingtalk.com/media/upload?access_token={access_token}&type=file"
    with open(file_path, 'rb') as f:
        files = {'media': (os.path.basename(file_path), f)}
        resp = requests.post(upload_url, files=files)
    upload_data = resp.json()
    logger.info(f"上传文件结果: {upload_data}")
    if upload_data.get("errcode") != 0:
        raise RuntimeError(f"上传文件失败: {upload_data}")
    return upload_data["media_id"]


def send_file_message(message: ChatbotMessage, file_path: str, file_name: str):
    """发送文件消息 — 上传 media 后通过 webhook 发送"""
    access_token = get_access_token()
    file_size = os.path.getsize(file_path)
    logger.info(f"发送文件: {file_name}, 大小: {file_size} 字节, 路径: {file_path}")

    # 上传文件获取 media_id
    upload_url = f"https://oapi.dingtalk.com/media/upload?access_token={access_token}&type=file"
    with open(file_path, 'rb') as f:
        files = {'media': (file_name, f, 'application/octet-stream')}
        resp = requests.post(upload_url, files=files)
    upload_data = resp.json()
    logger.info(f"上传文件结果: {upload_data}")

    if upload_data.get("errcode") != 0:
        raise RuntimeError(f"上传文件失败: {upload_data}")

    media_id = upload_data["media_id"]

    session_webhook = getattr(message, 'session_webhook', None)
    if not session_webhook:
        logger.error("无法获取session_webhook")
        return {"error": "no session_webhook"}

    # 确定文件类型
    ext = os.path.splitext(file_name)[1].lower().lstrip(".")
    file_type_map = {
        "pptx": "pptx", "ppt": "pptx",
        "xlsx": "xlsx", "xls": "xlsx",
        "docx": "docx", "doc": "docx",
        "pdf": "pdf",
        "zip": "zip", "rar": "zip",
        "txt": "txt", "csv": "csv",
        "png": "png", "jpg": "jpg", "jpeg": "jpg", "gif": "gif",
    }
    file_type = file_type_map.get(ext, ext)

    # 发送文件消息
    body = {
        "msgtype": "file",
        "file": {
            "mediaId": media_id,
            "fileName": file_name,
            "fileType": file_type
        },
        "at": {
            "atUserIds": [message.sender_staff_id]
        }
    }

    headers = {'Content-Type': 'application/json', 'Accept': '*/*'}
    resp = requests.post(session_webhook, json=body, headers=headers)
    result = resp.json()
    logger.info(f"发送文件结果: {result}")

    # 如果文件消息发送失败，回退到文本方式
    if result.get("errcode") != 0:
        logger.warning(f"文件消息发送失败，尝试文本回退")
        try:
            # 读取 Excel 内容并转为文本表格
            if ext in ('xlsx', 'xls'):
                import openpyxl
                wb = openpyxl.load_workbook(file_path)
                text_parts = [f"📄 {file_name}\n"]
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    text_parts.append(f"\n【{sheet_name}】")
                    for row in ws.iter_rows(values_only=True):
                        row_text = " | ".join([str(c) if c else "" for c in row])
                        text_parts.append(row_text)
                text_content = "\n".join(text_parts)
            else:
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    text_content = f"📄 {file_name}\n\n{f.read()[:4000]}"

            fallback_body = {
                "msgtype": "markdown",
                "markdown": {
                    "title": f"📄 {file_name}",
                    "text": f"```\n{text_content[:4000]}\n```"
                },
                "at": {"atUserIds": [message.sender_staff_id]}
            }
            resp2 = requests.post(session_webhook, json=fallback_body, headers=headers)
            logger.info(f"文本回退发送结果: {resp2.json()}")
        except Exception as e:
            logger.error(f"文本回退也失败: {e}")

    return result


def send_image_message(message: ChatbotMessage, image_path: str):
    """发送图片消息 - 尝试多种方式确保成功"""
    logger.info(f"开始发送图片: {image_path}")
    logger.info(f"文件是否存在: {os.path.exists(image_path)}")
    logger.info(f"文件大小: {os.path.getsize(image_path) if os.path.exists(image_path) else 'N/A'}")

    access_token = get_access_token()

    # 方式1: 上传图片获取 mediaId
    upload_url = f"https://oapi.dingtalk.com/media/upload?access_token={access_token}&type=image"
    with open(image_path, 'rb') as f:
        files = {'media': (os.path.basename(image_path), f, 'image/png')}
        resp = requests.post(upload_url, files=files)
    upload_data = resp.json()
    logger.info(f"上传图片结果: {upload_data}")

    if upload_data.get("errcode") != 0:
        logger.warning(f"图片上传失败，使用文件方式发送")
        send_file_message(message, image_path, os.path.basename(image_path))
        return

    media_id = upload_data["media_id"]

    # 方式2: 获取图片 URL（用于 picURL 方式）
    try:
        get_url = f"https://oapi.dingtalk.com/media/get?access_token={access_token}&media_id={media_id}"
        resp = requests.get(get_url, allow_redirects=False)
        pic_url = resp.headers.get('Location', '')
        logger.info(f"获取图片URL: {pic_url}")
    except Exception as e:
        logger.warning(f"获取图片URL失败: {e}")
        pic_url = ""

    session_webhook = getattr(message, 'session_webhook', None)
    if not session_webhook:
        logger.error("无法获取session_webhook")
        send_file_message(message, image_path, os.path.basename(image_path))
        return

    # 方式3: 尝试用 picURL 发送
    if pic_url:
        body = {
            "msgtype": "image",
            "image": {
                "picURL": pic_url
            }
        }
        headers = {'Content-Type': 'application/json', 'Accept': '*/*'}
        resp = requests.post(session_webhook, json=body, headers=headers)
        result = resp.json()
        logger.info(f"picURL 方式发送结果: {result}")
        if result.get("errcode") == 0 or not result.get("errcode"):
            return

    # 方式4: 尝试用 mediaId 发送
    body = {
        "msgtype": "image",
        "image": {
            "mediaId": media_id
        }
    }
    headers = {'Content-Type': 'application/json', 'Accept': '*/*'}
    resp = requests.post(session_webhook, json=body, headers=headers)
    result = resp.json()
    logger.info(f"mediaId 方式发送结果: {result}")
    if result.get("errcode") == 0 or not result.get("errcode"):
        return

    # 方式5: 所有方式都失败，使用文件方式
    logger.warning("所有图片发送方式都失败，使用文件方式")
    send_file_message(message, image_path, os.path.basename(image_path))


# 消息去重缓存
_processed_messages = {}

# 巡检补充照片待处理状态
# 格式: {user_id: {"record_id": str, "expire_time": float}}
_pending_photo_requests = {}


async def _archive_message_to_kb(message: ChatbotMessage, text_content: str,
                                 school_config: SchoolConfig,
                                 source_type: str = "text",
                                 file_name: str = "",
                                 file_path: str = ""):
    """将消息自动存入知识库"""
    try:
        # 过滤掉申请、审批相关的消息
        skip_keywords = [
            "申请查看权限", "申请查询", "申请访问", "申请权限",
            "同意 approval_", "拒绝 approval_", "审批",
            "已通知申请人", "审批通过", "审批拒绝"
        ]
        for keyword in skip_keywords:
            if keyword in text_content:
                logger.info(f"跳过存档消息（包含关键词: {keyword}）")
                return

        kb = get_knowledge_base(school_config.knowledge_dir, school_config.corp_id)
        msg_id = getattr(message, 'message_id', None) or getattr(message, 'msg_id', None)
        if not msg_id:
            msg_id = hashlib.md5(f"{message.sender_id}_{time.time()}".encode()).hexdigest()[:16]

        await kb.add_message(
            text=text_content,
            source_type=source_type,
            source_id=msg_id,
            sender_id=message.sender_staff_id or message.sender_id or "",
            sender_nick=message.sender_nick or "",
            conversation_id=message.conversation_id or "",
            message_type=message.message_type or "",
            file_name=file_name,
            file_path=file_path,
        )
    except Exception as e:
        logger.error(f"消息存档失败: {e}")


def _auto_add_user(user_id: str, user_name: str, corp_id: str):
    """自动添加新用户到用户列表"""
    if not user_id or not corp_id:
        return

    try:
        # 获取知识库目录
        from web.config import KNOWLEDGE_DIR
        users_file = os.path.join(KNOWLEDGE_DIR, corp_id, "structured", "users.json")

        # 加载现有用户
        users = []
        if os.path.exists(users_file):
            with open(users_file, 'r', encoding='utf-8') as f:
                users = json.load(f)

        # 检查用户是否已存在
        for user in users:
            if user.get("user_id") == user_id:
                # 用户已存在，更新最后活跃时间（可选）
                return

        # 添加新用户
        new_user = {
            "user_id": user_id,
            "name": user_name,
            "role": "teacher",  # 默认角色为教师
            "department": "",
            "manager_id": "",
            "permissions": ["public", "internal"],  # 默认权限
            "metadata": {
                "auto_created": True,
                "created_at": datetime.now().isoformat(),
            }
        }
        users.append(new_user)

        # 保存用户列表
        os.makedirs(os.path.dirname(users_file), exist_ok=True)
        with open(users_file, 'w', encoding='utf-8') as f:
            json.dump(users, f, ensure_ascii=False, indent=2)

        logger.info(f"自动添加新用户: {user_name} ({user_id})")
    except Exception as e:
        logger.error(f"自动添加用户失败: {e}")


async def _handle_image_message(message: ChatbotMessage, school_config: SchoolConfig) -> str:
    """处理图片消息：下载图片、OCR提取文字、存入知识库"""
    try:
        image_content = message.image_content
        if not image_content or not image_content.download_code:
            return "收到图片，但无法获取下载链接。"

        # 下载图片
        files_dir = os.path.join(school_config.knowledge_dir, "files", get_today_dir())
        file_path = await download_media_from_dingtalk(
            image_content.download_code, files_dir
        )

        if not file_path:
            return "收到图片，但下载失败。"

        # OCR提取文字
        extracted_text, file_type = await extract_text_from_file(file_path)
        if extracted_text and extracted_text.strip():
            # 检测是否为结构化数据
            data_type = detect_data_type(extracted_text)
            if data_type == "schedule":
                entries = await ScheduleParser.parse_from_text(extracted_text)
                if entries:
                    kb = get_knowledge_base(school_config.knowledge_dir, school_config.corp_id)
                    existing = kb.get_structured_data("schedules")
                    existing.extend([e.__dict__ for e in entries])
                    kb.save_structured_data("schedules", existing)

            # 存入知识库
            await _archive_message_to_kb(
                message, extracted_text, school_config,
                source_type="image", file_path=file_path
            )
            return f"图片已识别并存入知识库。\n识别内容预览：\n{extracted_text[:500]}{'...' if len(extracted_text) > 500 else ''}"
        else:
            await _archive_message_to_kb(
                message, "[图片内容无法识别]", school_config,
                source_type="image", file_path=file_path
            )
            return "图片已保存，但未能识别出文字内容。"

    except Exception as e:
        logger.error(f"处理图片消息失败: {e}", exc_info=True)
        return f"处理图片时出现错误：{str(e)}"


async def _handle_file_message(message: ChatbotMessage, school_config: SchoolConfig) -> Tuple[str, str]:
    """处理文件/富文本消息：下载文件/图片、提取文字、存入知识库

    返回:
        (回复消息, 文件路径)
    """
    try:
        # 尝试从不同位置获取文件/图片信息
        text_parts = []
        download_code = ""
        file_name = ""
        file_path = ""
        is_image = False

        # 调试：打印完整消息结构
        logger.info(f"=== 文件消息调试 ===")
        logger.info(f"message_type: {message.message_type}")
        logger.info(f"extensions: {getattr(message, 'extensions', None)}")
        logger.info(f"rich_text_content: {getattr(message, 'rich_text_content', None)}")

        # 方式0: 处理纯文件消息（file类型）
        if message.message_type == "file":
            # 文件消息的 downloadCode 在 extensions 中
            if message.extensions:
                # 尝试从不同路径获取 downloadCode
                download_code = (
                    message.extensions.get("downloadCode", "") or
                    message.extensions.get("content", {}).get("downloadCode", "") or
                    message.extensions.get("content", {}).get("fileDownloadCode", "")
                )
                file_name = (
                    message.extensions.get("fileName", "") or
                    message.extensions.get("content", {}).get("fileName", "") or
                    message.extensions.get("content", {}).get("name", "unknown")
                )
                logger.info(f"文件消息: downloadCode={download_code}, fileName={file_name}")

        # 方式1: 从 extensions.content.richText 获取（richText类型消息）
        if not download_code and message.extensions and "content" in message.extensions:
            content = message.extensions.get("content", {})
            rich_text_list = content.get("richText", [])
            logger.info(f"从extensions.content.richText获取: {len(rich_text_list)} 项")
            for item in rich_text_list:
                if isinstance(item, dict):
                    if item.get("type") == "picture" and not download_code:
                        download_code = item.get("downloadCode", "")
                        is_image = True
                        logger.info(f"找到图片downloadCode")
                    elif item.get("type") == "file" and not download_code:
                        download_code = item.get("downloadCode", "")
                        file_name = item.get("fileName", "unknown")
                    elif item.get("type") == "text":
                        text = item.get("text", "").strip()
                        if text:
                            text_parts.append(text)

        # 方式2: 从 rich_text_content 获取
        if not download_code:
            rich_text = message.rich_text_content
            if rich_text and rich_text.rich_text_list:
                for item in rich_text.rich_text_list:
                    if isinstance(item, dict):
                        if item.get("type") == "picture" and not download_code:
                            download_code = item.get("downloadCode", "")
                            is_image = True
                        elif item.get("type") == "file" and not download_code:
                            download_code = item.get("downloadCode", "")
                            file_name = item.get("fileName", "unknown")
                        elif item.get("type") == "text":
                            text = item.get("text", "").strip()
                            if text:
                                text_parts.append(text)

        if download_code:
            files_dir = os.path.join(school_config.knowledge_dir, "files", get_today_dir())
            if is_image:
                file_name = f"image_{hashlib.md5(download_code.encode()).hexdigest()[:8]}.jpg"
            file_path = await download_media_from_dingtalk(
                download_code, files_dir, file_name
            )

            if file_path:
                extracted_text, file_type = await extract_text_from_file(file_path)
                logger.info(f"文件提取结果: type={file_type}, text_len={len(extracted_text) if extracted_text else 0}")

                if extracted_text and extracted_text.strip():
                    # 如果有文本部分，合并
                    if text_parts:
                        combined = "\n".join(text_parts) + "\n" + extracted_text
                    else:
                        combined = extracted_text

                    # 检测结构化数据
                    data_type = detect_data_type(combined)
                    if data_type == "schedule":
                        entries = await ScheduleParser.parse_from_text(combined)
                        if entries:
                            kb = get_knowledge_base(school_config.knowledge_dir, school_config.corp_id)
                            existing = kb.get_structured_data("schedules")
                            existing.extend([e.__dict__ for e in entries])
                            kb.save_structured_data("schedules", existing)
                    elif data_type == "exam":
                        entries = await ExamParser.parse_from_text(combined)
                        if entries:
                            kb = get_knowledge_base(school_config.knowledge_dir, school_config.corp_id)
                            existing = kb.get_structured_data("exams")
                            existing.extend([e.__dict__ for e in entries])
                            kb.save_structured_data("exams", existing)
                    elif data_type == "contact":
                        entries = await ContactParser.parse_from_text(combined)
                        if entries:
                            kb = get_knowledge_base(school_config.knowledge_dir, school_config.corp_id)
                            existing = kb.get_structured_data("contacts")
                            existing.extend([e.__dict__ for e in entries])
                            kb.save_structured_data("contacts", existing)

                    # 存入知识库
                    source_type = "image" if is_image else "file"
                    await _archive_message_to_kb(
                        message, combined, school_config,
                        source_type=source_type, file_name=file_name, file_path=file_path
                    )
                    return f"{'图片' if is_image else '文件'}已识别并存入知识库。\n识别内容：\n{combined[:500]}{'...' if len(combined) > 500 else ''}", file_path
                else:
                    # 没有提取到文字，但有文本部分
                    if text_parts:
                        combined = "\n".join(text_parts)
                        await _archive_message_to_kb(message, combined, school_config)
                        return f"消息已存入知识库。", file_path

                    await _archive_message_to_kb(
                        message, f"[{'图片' if is_image else '文件'}：{file_name}]", school_config,
                        source_type="image" if is_image else "file", file_name=file_name, file_path=file_path
                    )
                    return f"{'图片' if is_image else '文件'}已保存，但未能识别出文字内容。", file_path

        # 如果只有文本部分
        if text_parts:
            combined_text = "\n".join(text_parts)
            await _archive_message_to_kb(message, combined_text, school_config)
            return f"消息已存入知识库。", file_path

        # 调试：打印最终状态
        logger.info(f"=== 解析结束 ===")
        logger.info(f"download_code: '{download_code}'")
        logger.info(f"file_name: '{file_name}'")
        logger.info(f"text_parts: {text_parts}")

        return "收到消息，但未能解析内容。可能是文件格式不支持或下载链接获取失败。", file_path

    except Exception as e:
        logger.error(f"处理消息失败: {e}", exc_info=True)
        return f"处理消息时出现错误：{str(e)}", ""


class SchoolBotHandler(ChatbotHandler):
    """学校智能助手消息处理器"""

    def _log_message(self, message: ChatbotMessage, text: str, status: str = "success",
                     skill_used: str = "", error_msg: str = "", processing_time_ms: int = 0,
                     kb_results_count: int = 0):
        """记录消息日志到 Web 数据库"""
        if not _web_log_enabled:
            return

        try:
            db = SessionLocal()
            try:
                msg_id = getattr(message, 'message_id', None) or getattr(message, 'msg_id', None)
                if not msg_id:
                    msg_id = hashlib.md5(f"{message.sender_id}_{time.time()}".encode()).hexdigest()[:16]

                # 检查是否已存在
                existing = db.query(MessageLog).filter(MessageLog.msg_id == msg_id).first()
                if existing:
                    # 更新现有记录
                    existing.status = status
                    existing.error_msg = error_msg or existing.error_msg
                    existing.processing_time_ms = processing_time_ms or existing.processing_time_ms
                    existing.skill_used = skill_used or existing.skill_used
                    existing.kb_results_count = kb_results_count or existing.kb_results_count
                else:
                    # 创建新记录
                    log = MessageLog(
                        msg_id=msg_id,
                        sender_id=message.sender_staff_id or message.sender_id or "",
                        sender_nick=message.sender_nick or "",
                        content=text[:2000] if text else "",
                        message_type=message.message_type or "text",
                        conversation_id=message.conversation_id or "",
                        corp_id=getattr(message, 'sender_corp_id', None) or "",
                        status=status,
                        error_msg=error_msg,
                        processing_time_ms=processing_time_ms,
                        skill_used=skill_used,
                        kb_results_count=kb_results_count,
                    )
                    db.add(log)

                db.commit()
            finally:
                db.close()
        except Exception as e:
            logger.error(f"记录消息日志失败: {e}")

    async def process(self, callback: dingtalk_stream.CallbackMessage):
        try:
            message = ChatbotMessage.from_dict(callback.data)
            logger.info(f"=== 收到消息 ===")
            logger.info(f"sender_nick: {message.sender_nick}")

            # 记录消息开始处理
            start_time = time.time()
            msg_text = ""
            if message.message_type == "text" and message.text:
                msg_text = message.text.content or ""
            self._log_message(message, msg_text, status="processing")

            # 获取corp_id，识别学校
            corp_id = getattr(message, 'sender_corp_id', None) or ""
            school_config = school_manager.get_school(corp_id)
            logger.info(f"学校: {school_config.name} ({corp_id})")

            # 提前获取用户信息（图片消息也需要）
            sender_nick = message.sender_nick or "老师"
            user_id = message.sender_staff_id or message.sender_id
            conversation_id = message.conversation_id or ""

            # 处理不同类型的消息
            logger.info(f"消息类型: {message.message_type}")

            # 调试：打印消息原始数据
            if message.message_type in ("picture", "richText", "file"):
                logger.info(f"image_content: {message.image_content}")
                logger.info(f"rich_text_content: {message.rich_text_content}")
                logger.info(f"extensions: {message.extensions}")

            if message.message_type == "picture":
                # 检查是否有待处理的巡检补充照片请求
                pending_photo = _pending_photo_requests.get(user_id)
                if pending_photo and pending_photo["expire_time"] > time.time():
                    # 有待处理的照片请求，传递给巡检技能
                    logger.info(f"处理巡检补充照片: user={user_id}, record={pending_photo['record_id']}")
                    try:
                        from agent.inspection.service import get_inspection_service
                        from agent.skills.inspection_skill import InspectionSkill

                        service = get_inspection_service()

                        # 下载图片
                        files_dir = os.path.join(school_config.knowledge_dir, "files", get_today_dir())
                        from agent.media_handler import download_media_from_dingtalk
                        image_content = message.image_content
                        if image_content and image_content.download_code:
                            file_path = await download_media_from_dingtalk(
                                image_content.download_code, files_dir
                            )
                            if file_path:
                                # 获取记录
                                record_id = pending_photo["record_id"]
                                if not record_id:
                                    # 没有指定记录ID，使用最近的记录
                                    records = service.list_records(inspector_id=user_id)
                                    if records:
                                        record = records[0]
                                        record_id = record.record_id
                                    else:
                                        self.reply_text("❌ 没有找到打卡记录", message)
                                        del _pending_photo_requests[user_id]
                                        return AckMessage.STATUS_OK, "OK"
                                else:
                                    record = service._records.get(record_id)

                                if record:
                                    # 保存照片到记录
                                    if record.photo_urls is None:
                                        record.photo_urls = []
                                    record.photo_urls.append(file_path)
                                    service._save_records()

                                    self.reply_text(
                                        f"✅ 照片已补充！\n\n"
                                        f"📋 记录ID：{record_id}\n"
                                        f"📷 累计照片：{len(record.photo_urls)} 张",
                                        message
                                    )
                                else:
                                    self.reply_text(f"❌ 未找到记录「{record_id}」", message)

                                # 清除待处理状态
                                del _pending_photo_requests[user_id]
                            else:
                                self.reply_text("⚠️ 图片下载失败，请重试", message)
                        else:
                            self.reply_text("⚠️ 无法获取图片下载链接", message)

                    except Exception as e:
                        logger.error(f"处理巡检照片失败: {e}", exc_info=True)
                        self.reply_text(f"⚠️ 处理照片失败：{str(e)}", message)
                        if user_id in _pending_photo_requests:
                            del _pending_photo_requests[user_id]

                    processing_time = int((time.time() - start_time) * 1000)
                    self._log_message(message, "[巡检补充照片]", status="success", processing_time_ms=processing_time)
                    return AckMessage.STATUS_OK, "OK"

                # 普通图片消息
                result = await _handle_image_message(message, school_config)
                self.reply_text(result, message)
                processing_time = int((time.time() - start_time) * 1000)
                self._log_message(message, "[图片消息]", status="success", processing_time_ms=processing_time)
                return AckMessage.STATUS_OK, "OK"

            elif message.message_type == "richText":
                # 富文本消息（可能包含文件）
                result, file_path = await _handle_file_message(message, school_config)
                self.reply_text(result, message)
                processing_time = int((time.time() - start_time) * 1000)
                self._log_message(message, "[富文本消息]", status="success", processing_time_ms=processing_time)
                return AckMessage.STATUS_OK, "OK"

            elif message.message_type == "file":
                # 文件消息
                result, file_path = await _handle_file_message(message, school_config)
                self.reply_text(result, message)
                # 检查是否是Excel文件，如果是，调用排课技能
                if message.extensions:
                    file_name = (
                        message.extensions.get("fileName", "") or
                        message.extensions.get("content", {}).get("fileName", "") or
                        message.extensions.get("content", {}).get("name", "")
                    )
                    if file_name and file_name.endswith(('.xlsx', '.xls')):
                        # 检测是否是排课相关的Excel
                        if '排课' in file_name or '课表' in file_name:
                            # 调用排课技能
                            skill_match = skill_registry.match("排课模板")
                            if skill_match and skill_match.confidence >= 0.7:
                                context = {
                                    'sender_nick': sender_nick,
                                    'user_id': user_id,
                                    'conversation_id': conversation_id,
                                    'corp_id': corp_id,
                                    'school_config': school_config,
                                    'message': message,
                                    '_file_path': file_path,
                                }
                                result = await skill_match.skill.execute("排课模板", context)
                                if result:
                                    self.reply_text(result, message)
                        # 检测是否是巡检相关的Excel
                        elif '巡检' in file_name or '点位' in file_name or '问题' in file_name:
                            from agent.inspection.service import get_inspection_service
                            service = get_inspection_service()
                            if '点位' in file_name:
                                imported, skipped, msg = service.import_points_from_excel(file_path)
                            else:
                                imported, skipped, msg = service.import_issues_from_excel(file_path)
                            self.reply_text(msg, message)
                processing_time = int((time.time() - start_time) * 1000)
                self._log_message(message, "[文件消息]", status="success", processing_time_ms=processing_time)
                return AckMessage.STATUS_OK, "OK"

            elif message.message_type != "text":
                # 其他类型消息，记录但不处理
                logger.info(f"忽略非文本消息类型: {message.message_type}")
                processing_time = int((time.time() - start_time) * 1000)
                self._log_message(message, f"[忽略: {message.message_type}]", status="skipped", processing_time_ms=processing_time)
                return AckMessage.STATUS_OK, "OK"

            # 处理文本消息
            text = message.text.content.strip() if message.text and message.text.content else ""
            if not text:
                return AckMessage.STATUS_OK, "OK"

            if message.is_in_at_list:
                text = text.replace(f"@{message.robot_code}", "").strip()

            logger.info(f"收到消息 [{sender_nick}]: {text}")

            # 自动添加新用户到用户列表
            _auto_add_user(user_id, sender_nick, corp_id)

            # 检查是否有超时的巡检记录（30分钟未活动自动签退）
            try:
                from agent.inspection.service import get_inspection_service
                inspection_service = get_inspection_service()
                # 查找该用户未签退的记录
                user_records = inspection_service.list_records(inspector_id=user_id)
                current_time = time.time()
                for record in user_records:
                    if record.check_out_time == 0:
                        # 计算距离打卡时间的间隔（分钟）
                        time_since_checkin = (current_time - record.check_in_time) / 60
                        # 如果超过30分钟未签退，自动签退
                        if time_since_checkin > 30:
                            inspection_service.check_out(
                                record_id=record.record_id,
                                check_results=[],
                            )
                            logger.info(f"自动签退: 用户 {user_nick}，记录 {record.record_id}（超过30分钟未活动）")
            except Exception as e:
                logger.error(f"自动签退检查失败: {e}")

            # 自动存入知识库（纯文本消息，带过滤）
            from agent.knowledge_base_v2 import should_skip_message
            if not should_skip_message(text):
                await _archive_message_to_kb(message, text, school_config)
            else:
                logger.debug(f"跳过无意义消息: {text[:50]}")

            # 消息去重
            message_id = getattr(message, 'message_id', None) or getattr(message, 'msg_id', None)
            if message_id:
                current_time = time.time()
                expired_keys = [k for k, v in _processed_messages.items() if current_time - v > 300]
                for k in expired_keys:
                    del _processed_messages[k]
                if message_id in _processed_messages:
                    return AckMessage.STATUS_OK, "OK"
                _processed_messages[message_id] = current_time

            # 检查是否有待确认的PPT任务
            pending_task = state_manager.get_pending_task(user_id, conversation_id, corp_id)
            if pending_task:
                try:
                    if self._handle_confirmation(text, pending_task, message, corp_id):
                        return AckMessage.STATUS_OK, "OK"
                except Exception as e:
                    logger.error(f"处理PPT确认失败: {e}", exc_info=True)
                    self.reply_text(f"处理确认时出错：{str(e)}", message)
                    return AckMessage.STATUS_OK, "OK"

            # ─── LLM 意图识别 + 技能系统匹配 ───
            # 首先使用 LLM 识别意图
            intent_context = {
                'sender_nick': sender_nick,
                'user_id': user_id,
                'conversation_id': conversation_id,
                'corp_id': corp_id,
                'current_time': datetime.now().isoformat(),
            }
            intent = await intent_router.classify(text, intent_context)
            logger.info(f"LLM意图识别: {intent.type}/{intent.action} (置信度: {intent.confidence:.2f}, 来源: {intent.raw_text[:50]})")

            # 根据意图类型分发到对应技能
            skill = None
            logger.info(f"意图类型: {intent.type}")

            if intent.type in ["inspection", "asset", "schedule", "ppt", "knowledge"]:
                # 使用意图路由器识别的类型来匹配技能
                skill_name_map = {
                    "inspection": "巡检管理",
                    "asset": "资产管理",
                    "schedule": "课表管理",
                    "ppt": "PPT生成",
                    "knowledge": "知识库",
                }
                skill_name = skill_name_map.get(intent.type)
                logger.info(f"技能名称: {skill_name}")

                if skill_name:
                    skill = skill_registry.get_skill(skill_name)
                    logger.info(f"找到技能: {skill is not None}")

                    if skill:
                        logger.info(f"LLM意图匹配到技能: {skill.name}")

            # 如果 LLM 没有匹配到，回退到原有规则匹配
            if not skill:
                skill_match_result = skill_registry.match(text)
                if skill_match_result and skill_match_result.confidence >= 0.7:
                    skill = skill_match_result.skill
                    logger.info(f"规则匹配到技能: {skill.name} (置信度: {skill_match_result.confidence:.2f})")

            if skill:
                logger.info(f"匹配到技能: {skill.name}")
                try:
                    # 获取用户角色
                    from agent.permission_manager import get_permission_manager
                    perm_manager = get_permission_manager(school_config.knowledge_dir, corp_id)
                    user_role = perm_manager.get_user_role(user_id)

                    context = {
                        'sender_nick': sender_nick,
                        'user_id': user_id,
                        'user_role': user_role,
                        'conversation_id': conversation_id,
                        'corp_id': corp_id,
                        'school_config': school_config,
                        'message': message,
                        'intent': intent,  # 传递 LLM 识别的意图
                    }

                    # 检查是否是巡检技能的"补充照片"操作，设置待处理状态
                    from agent.skills.inspection_skill import InspectionSkill
                    if isinstance(skill, InspectionSkill) and "补充照片" in text:
                        # 解析记录ID
                        record_id = ""
                        for prefix in ["补充照片"]:
                            if prefix in text.lower():
                                parts = text.lower().split(prefix, 1)
                                if len(parts) > 1:
                                    record_id = parts[1].strip()
                                    break

                        # 设置待处理照片状态（5分钟有效期）
                        _pending_photo_requests[user_id] = {
                            "record_id": record_id,
                            "expire_time": time.time() + 300,
                            "corp_id": corp_id,
                        }
                        logger.info(f"设置待补充照片状态: user={user_id}, record={record_id}")

                    result = await skill.execute(text, context)
                    if result:
                        self.reply_text(result, message)

                        # 检查巡检打卡后是否需要拍照
                        from agent.skills.inspection_skill import InspectionSkill
                        if isinstance(skill, InspectionSkill) and ("请补充现场照片" in result or "该点位要求拍照" in result):
                            # 从上下文获取记录ID
                            record_id = context.get("_last_checkin_record_id", "")
                            # 设置待处理照片状态（5分钟有效期）
                            _pending_photo_requests[user_id] = {
                                "record_id": record_id,
                                "expire_time": time.time() + 300,
                                "corp_id": corp_id,
                            }
                            logger.info(f"巡检打卡后设置待补充照片状态: user={user_id}, record={record_id}")

                        # 如果技能设置了待发送文件，通过 session_webhook 发送
                        file_to_send = context.get("_file_to_send")
                        file_type = context.get("_file_type", "file")
                        file_name = context.get("_file_name", "")
                        logger.info(f"技能返回文件检查: file_to_send={file_to_send}, file_type={file_type}, file_name={file_name}")
                        if file_to_send and os.path.exists(file_to_send):
                            try:
                                if file_type == "image":
                                    send_image_message(message, file_to_send)
                                else:
                                    send_file_message(message, file_to_send, file_name)
                                logger.info(f"技能文件已发送: {file_name}")
                            except Exception as e:
                                logger.error(f"技能文件发送失败: {e}", exc_info=True)
                                # 发送失败时通知用户
                                self.reply_text(f"⚠️ 文件生成成功但发送失败：{str(e)}\n\n文件已保存到：{file_to_send}", message)

                        # 如果是调课操作，发送通知到指定用户
                        swap_notify_to = context.get("_swap_notify_to")
                        swap_notify_msg = context.get("_swap_notify_msg")
                        if swap_notify_to and swap_notify_msg:
                            try:
                                # 通过 OpenAPI 发送消息给指定用户
                                from dingtalk.bot import reply_text as dingtalk_reply
                                # 使用绑定的钉钉用户ID发送消息
                                logger.info(f"调课通知目标: {swap_notify_to}")
                                logger.info(f"调课通知内容: {swap_notify_msg[:100]}...")

                                # 尝试向目标用户发送消息
                                # 注意：这里需要目标用户在同一个会话中，或者使用其他方式发送
                                # 暂时通过当前会话发送通知
                                self.reply_text(swap_notify_msg, message)
                                logger.info("调课通知已发送")
                            except Exception as e:
                                logger.error(f"调课通知发送失败: {e}")

                        # 兼容旧的通知方式
                        swap_notification = context.get("_swap_notification")
                        if swap_notification:
                            try:
                                self.reply_text(swap_notification, message)
                                logger.info("调课通知已发送")
                            except Exception as e:
                                logger.error(f"调课通知发送失败: {e}")
                        else:
                            logger.info(f"没有待发送的文件或文件不存在: {file_to_send}")
                        processing_time = int((time.time() - start_time) * 1000)
                        self._log_message(message, text, status="success", skill_used=skill.name, processing_time_ms=processing_time)
                        return AckMessage.STATUS_OK, "OK"
                except Exception as e:
                    logger.error(f"技能 {skill.name} 执行失败: {e}", exc_info=True)
                    self.reply_text(f"技能执行出错：{str(e)}", message)
                    processing_time = int((time.time() - start_time) * 1000)
                    self._log_message(message, text, status="error", skill_used=skill.name, error_msg=str(e), processing_time_ms=processing_time)
                    return AckMessage.STATUS_OK, "OK"

            # ─── PPT 队列查询 ───
            queue_keywords = ["PPT队列", "队列状态", "排队状态", "PPT进度", "生成进度"]
            if any(keyword in text for keyword in queue_keywords):
                from agent.ppt_task_manager import get_ppt_task_manager
                task_manager = get_ppt_task_manager()
                queue_status = task_manager.get_all_queue_status()

                # 构建状态消息
                concurrency = queue_status.get('concurrency', {})
                queue_list = queue_status.get('queue', [])
                stats = queue_status.get('stats', {})
                avg_time = queue_status.get('avg_generation_time', 120)

                status_msg = f"📊 PPT生成队列状态\n\n"
                status_msg += f"⚙️ 并发配置：{concurrency.get('max', 5)} 个线程\n"
                status_msg += f"🚀 正在生成：{concurrency.get('running', 0)} 个\n"
                status_msg += f"⏳ 排队等待：{len(queue_list)} 个\n"
                status_msg += f"📈 平均生成时间：{avg_time}秒\n\n"

                if queue_list:
                    status_msg += "📋 排队列表：\n"
                    for item in queue_list[:5]:  # 最多显示5个
                        status_msg += f"  {item['position']}. {item['topic'][:20]} ({item['estimated_wait_display']})\n"
                    if len(queue_list) > 5:
                        status_msg += f"  ... 还有 {len(queue_list) - 5} 个任务\n"
                else:
                    status_msg += "✅ 当前没有排队任务\n"

                # 统计信息
                status_msg += f"\n📈 历史统计：\n"
                status_msg += f"  总任务：{stats.get('total', 0)}\n"
                status_msg += f"  已完成：{stats.get('completed', 0)}\n"
                status_msg += f"  失败：{stats.get('failed', 0)}\n"

                self.reply_text(status_msg, message)
                processing_time = int((time.time() - start_time) * 1000)
                self._log_message(message, text, status="success", skill_used="PPT队列查询", processing_time_ms=processing_time)
                return AckMessage.STATUS_OK, "OK"

            # ─── PPT 生成请求（统一走 ppt-master，异步执行）───
            # 使用 LLM 识别的意图或原有关键词匹配
            is_ppt = False
            ppt_info = {}

            if intent.type == "ppt":
                # 使用 LLM 识别的意图（支持多种 action）
                is_ppt = True
                ppt_info = {
                    'topic': intent.params.get("topic", "演示文稿"),
                    'subject': intent.params.get("subject", ""),
                    'grade': intent.params.get("grade", ""),
                    'difficulty': intent.params.get("difficulty", "中等"),
                    'content_type': '课件',
                    'chapter': '',
                    'page_count': intent.params.get("page_count"),
                }
                logger.info(f"使用 LLM 意图触发 PPT 生成: {ppt_info}")
            elif is_ppt_request(text):
                # 降级到原有关键词匹配
                is_ppt = True
                ppt_info = parse_education_info(text) if is_education_request(text) else {
                    'topic': text, 'subject': '', 'grade': '', 'difficulty': '中等', 'content_type': '课件', 'chapter': ''
                }
                # 清理主题
                topic = ppt_info['topic']
                for kw in PPT_KEYWORDS + ["生成", "制作", "帮我", "请", "一个"]:
                    topic = topic.replace(kw, "")
                topic = topic.strip()
                if not topic:
                    topic = "演示文稿"
                ppt_info['topic'] = topic
                logger.info(f"使用关键词匹配触发 PPT 生成: {ppt_info}")

            if is_ppt:
                logger.info("检测到PPT生成请求")
                info = ppt_info
                topic = info['topic']

                self.reply_text(f"✅ 已收到「{topic}」PPT生成请求，正在后台生成大纲，请稍候...", message)

                # 异步执行大纲生成（不阻塞事件循环）
                import time as time_module
                task_id = f"outline_{user_id}_{int(time_module.time() * 1000)}"

                def generate_outline_task():
                    """大纲生成任务（在线程池中执行）"""
                    try:
                        # 搜索教材内容
                        search_context = ""
                        try:
                            import asyncio
                            from agent.web_search import search_textbook_content
                            search_context = asyncio.run(search_textbook_content(
                                topic=topic,
                                subject=info.get('subject', ''),
                                grade=info.get('grade', ''),
                            ))
                            logger.info(f"教材搜索完成，获取 {len(search_context)} 字符")
                        except Exception as e:
                            logger.warning(f"教材搜索失败: {e}")

                        outline = generate_ppt_outline_markdown(
                            topic=topic,
                            subject=info.get('subject', ''),
                            grade=info.get('grade', ''),
                            search_context=search_context,
                        )

                        # 保存大纲到状态
                        state_manager.create_task(
                            user_id=user_id,
                            conversation_id=conversation_id,
                            task_type=TaskType.PPT_GENERATION,
                            original_request=text,
                            outline_markdown=outline,
                            outline_data=info,
                            corp_id=corp_id,
                        )

                        # 发送大纲给用户
                        confirm_message = f"""📋 为您生成的PPT大纲：

{outline}

---
请回复：
• "确认" - 使用此大纲生成PPT
• "修改：[您的修改意见]" - 告诉我需要如何调整
• "取消" - 取消本次生成"""

                        self.reply_text(confirm_message, message)
                        logger.info(f"大纲生成完成并已发送: {topic}")

                    except Exception as e:
                        logger.error(f"生成大纲失败: {e}", exc_info=True)
                        self.reply_text(f"抱歉，生成大纲时出现错误：{str(e)}", message)

                # 提交到线程池执行
                from agent.ppt_task_manager import get_ppt_task_manager
                task_manager = get_ppt_task_manager()
                loop = asyncio.get_event_loop()
                loop.run_in_executor(task_manager.executor, generate_outline_task)

                processing_time = int((time.time() - start_time) * 1000)
                self._log_message(message, text, status="success", skill_used="PPT生成", processing_time_ms=processing_time)
                return AckMessage.STATUS_OK, "OK"

            # ─── 搜索请求 ───
            elif is_search_request(text):
                logger.info("检测到搜索请求")
                self.reply_text("正在为您搜索，请稍候...", message)
                try:
                    query = extract_search_query(text)
                    search_result = search_and_summarize(query)
                    if len(search_result) > 4000:
                        search_result = search_result[:4000] + "\n\n...(结果过长已截断)"
                    self.reply_text(search_result, message)
                    processing_time = int((time.time() - start_time) * 1000)
                    self._log_message(message, text, status="success", skill_used="网络搜索", processing_time_ms=processing_time)
                except Exception as e:
                    logger.error(f"搜索失败: {e}", exc_info=True)
                    self.reply_text(f"抱歉，搜索时出现错误：{str(e)}", message)
                    processing_time = int((time.time() - start_time) * 1000)
                    self._log_message(message, text, status="error", skill_used="网络搜索", error_msg=str(e), processing_time_ms=processing_time)
                return AckMessage.STATUS_OK, "OK"

            # ─── 资源搜索 ───
            elif is_resource_request(text):
                # 检查是否是资产管理的模板请求（优先级高于资源搜索）
                asset_template_keywords = ["资产模板", "资产录入模板", "资产盘点模板", "批量导入模板"]
                is_asset_template = any(kw in text for kw in asset_template_keywords)

                if is_asset_template:
                    # 资产管理的模板请求，不走资源搜索
                    logger.info("检测到资产管理模板请求，跳过资源搜索")
                    # 继续后面的技能匹配逻辑
                else:
                    logger.info("检测到资源搜索请求")
                if "习题" in text or "试题" in text or "练习" in text:
                    self.reply_text("正在为您搜索习题资源...", message)
                    try:
                        topic = text
                        for kw in ["搜索", "查找", "找", "习题", "试题", "练习题", "帮我"]:
                            topic = topic.replace(kw, "")
                        result = search_exam_questions(topic.strip())
                        self.reply_text(result, message)
                        processing_time = int((time.time() - start_time) * 1000)
                        self._log_message(message, text, status="success", skill_used="习题搜索", processing_time_ms=processing_time)
                    except Exception as e:
                        self.reply_text(f"搜索失败：{str(e)}", message)
                        processing_time = int((time.time() - start_time) * 1000)
                        self._log_message(message, text, status="error", skill_used="习题搜索", error_msg=str(e), processing_time_ms=processing_time)
                elif "素材" in text or "图片" in text or "视频" in text or "动画" in text:
                    self.reply_text("正在为您搜索教学素材...", message)
                    try:
                        topic = text
                        material_type = "图片"
                        for kw in ["搜索", "查找", "找", "帮我"]:
                            topic = topic.replace(kw, "")
                        if "视频" in text:
                            material_type = "视频"
                        elif "动画" in text:
                            material_type = "动画"
                        for kw in ["素材", "图片", "视频", "动画"]:
                            topic = topic.replace(kw, "")
                        result = search_materials(topic.strip(), material_type)
                        self.reply_text(result, message)
                        processing_time = int((time.time() - start_time) * 1000)
                        self._log_message(message, text, status="success", skill_used="素材搜索", processing_time_ms=processing_time)
                    except Exception as e:
                        self.reply_text(f"搜索失败：{str(e)}", message)
                        processing_time = int((time.time() - start_time) * 1000)
                        self._log_message(message, text, status="error", skill_used="素材搜索", error_msg=str(e), processing_time_ms=processing_time)
                else:
                    self.reply_text("正在为您搜索教学资源...", message)
                    try:
                        topic = text
                        for kw in ["搜索", "查找", "找", "资源", "模板", "帮我"]:
                            topic = topic.replace(kw, "")
                        result = search_teaching_resources(topic.strip())
                        self.reply_text(result, message)
                        processing_time = int((time.time() - start_time) * 1000)
                        self._log_message(message, text, status="success", skill_used="资源搜索", processing_time_ms=processing_time)
                    except Exception as e:
                        self.reply_text(f"搜索失败：{str(e)}", message)
                        processing_time = int((time.time() - start_time) * 1000)
                        self._log_message(message, text, status="error", skill_used="资源搜索", error_msg=str(e), processing_time_ms=processing_time)
                # 如果是资产管理模板请求，不返回，继续后面的技能匹配
                if not is_asset_template:
                    return AckMessage.STATUS_OK, "OK"

            # ─── 新闻查询 ───
            elif is_news_request(text):
                logger.info("检测到新闻查询请求")
                self.reply_text("正在为您查询最新信息...", message)
                try:
                    query = text
                    # 清理搜索关键词，提取核心查询
                    for kw in ["查询", "查找", "帮我查", "查一下", "找一下", "搜一下",
                                "最新", "最近", "今日", "今天", "热点", "新闻",
                                "请", "帮我", "一下", "关于", "的"]:
                        query = query.replace(kw, "")
                    # 清理多余空格
                    query = ' '.join(query.split()).strip()
                    # 如果清理后为空，使用原文
                    if not query:
                        query = text
                    search_query = f"{query} 最新消息"
                    result = search_and_summarize(search_query)
                    if len(result) > 4000:
                        result = result[:4000] + "\n\n...(结果过长已截断)"
                    self.reply_text(result, message)
                    processing_time = int((time.time() - start_time) * 1000)
                    self._log_message(message, text, status="success", skill_used="新闻查询", processing_time_ms=processing_time)
                except Exception as e:
                    logger.error(f"查询失败: {e}", exc_info=True)
                    self.reply_text(f"抱歉，查询时出现错误：{str(e)}", message)
                    processing_time = int((time.time() - start_time) * 1000)
                    self._log_message(message, text, status="error", skill_used="新闻查询", error_msg=str(e), processing_time_ms=processing_time)
                return AckMessage.STATUS_OK, "OK"

            # ─── 普通对话（知识库检索 + 搜索增强）───
            else:
                # 先从知识库检索相关内容
                from agent.permission_manager import get_permission_manager
                perm_manager = get_permission_manager(school_config.knowledge_dir, corp_id)
                user_role = perm_manager.get_user_role(user_id)

                kb = get_knowledge_base(school_config.knowledge_dir, school_config.corp_id)
                search_result = await kb.search(text, top_k=10, user_id=user_id, user_nick=sender_nick, user_role=user_role)
                kb_results = search_result.get("results", []) if isinstance(search_result, dict) else search_result
                permission_info = search_result.get("permission_info", {})

                knowledge_context = ""
                source_files = []  # 收集来源文档
                if kb_results:
                    # 过滤低分结果，只保留有实质相关性的
                    MIN_SCORE = 0.35  # 优化：略微降低阈值，提高召回率
                    high_quality_results = [r for r in kb_results if r.score >= MIN_SCORE]

                    # 去重：相同内容只保留分数最高的
                    seen_hashes = set()
                    deduped_results = []
                    for r in high_quality_results:
                        chunk_hash = r.chunk.content_hash if hasattr(r.chunk, 'content_hash') else r.chunk.text[:50]
                        if chunk_hash not in seen_hashes:
                            seen_hashes.add(chunk_hash)
                            deduped_results.append(r)

                    # 最多传 5 条给 LLM，提供更丰富的上下文
                    top_results = deduped_results[:5]

                    if top_results:
                        knowledge_context = "以下是从学校知识库中检索到的相关信息：\n\n"
                        # 收集高相关度结果的来源（分数>0.6）
                        SOURCE_THRESHOLD = 0.6
                        for result in top_results:
                            if result.score >= SOURCE_THRESHOLD and result.chunk.file_name:
                                if result.chunk.file_name not in source_files:
                                    source_files.append(result.chunk.file_name)

                        for i, result in enumerate(top_results, 1):
                            chunk = result.chunk
                            score = result.score
                            # 只传文本内容，不传技术细节
                            knowledge_context += f"{i}. [相关度:{score:.2f}] {chunk.text}\n"
                            if chunk.sender_nick:
                                knowledge_context += f"   （来自：{chunk.sender_nick}）\n"
                            knowledge_context += "\n"

                        logger.info(f"知识库检索到 {len(kb_results)} 条，过滤后 {len(high_quality_results)} 条，去重后 {len(deduped_results)} 条，传给LLM {len(top_results)} 条")
                    else:
                        logger.info(f"知识库检索到 {len(kb_results)} 条，但全部低于分数门槛 {MIN_SCORE}")
                else:
                    logger.info("知识库检索无结果")

                # 添加权限提示
                if permission_info.get("has_restricted"):
                    restricted_count = permission_info.get("restricted_count", 0)
                    knowledge_context += f"\n⚠️ 注意：有 {restricted_count} 条受限内容未显示。"
                    knowledge_context += f"\n如需查看受限内容，请回复「申请查询 {text}」发起权限申请。\n"

                # 判断是否需要网络搜索
                need_search = False
                search_keywords = ["是什么", "怎么", "为什么", "哪些", "介绍", "解释", "说明", "推荐",
                                  "最新", "现在", "目前", "今年", "2026", "2025", "新闻", "动态"]
                for keyword in search_keywords:
                    if keyword in text:
                        need_search = True
                        break
                if not need_search and len(text) > 15:
                    need_search = True

                # 提取意图信息
                intent = search_result.get("intent") if isinstance(search_result, dict) else None

                # 使用知识库增强的回答（支持意图驱动）
                reply = await chat_with_knowledge(
                    user_message=text,
                    knowledge_context=knowledge_context,
                    need_web_search=need_search,
                    intent=intent,
                )

                # 添加来源文档信息
                if source_files:
                    reply += "\n\n📚 以上信息来自：\n"
                    for file_name in source_files:
                        reply += f"• {file_name}\n"

                if len(reply) > 4000:
                    reply = reply[:4000] + "\n\n...(内容过长已截断)"
                self.reply_text(reply, message)
                processing_time = int((time.time() - start_time) * 1000)
                kb_count = len(kb_results) if kb_results else 0
                self._log_message(message, text, status="success", skill_used="AI对话", processing_time_ms=processing_time, kb_results_count=kb_count)
                return AckMessage.STATUS_OK, "OK"

        except Exception as e:
            logger.error(f"处理消息异常: {e}", exc_info=True)
            processing_time = int((time.time() - start_time) * 1000) if 'start_time' in locals() else 0
            self._log_message(message, text if 'text' in locals() else "", status="error", error_msg=str(e), processing_time_ms=processing_time)
            return AckMessage.STATUS_OK, "OK"

    def _handle_confirmation(self, text: str, task_state, message: ChatbotMessage,
                             corp_id: str = "") -> bool:
        """处理大纲确认/修改/取消（兼容旧的 PENDING_TEMPLATE 状态）"""
        user_id = message.sender_staff_id or message.sender_id
        conversation_id = message.conversation_id or ""
        sender_nick = message.sender_nick or "老师"

        # 如果是旧的 PENDING_TEMPLATE 状态，强制转为 PENDING_OUTLINE
        # 这样用户回复"确认"时直接走新流程，不再走模板选择
        if task_state.status == TaskStatus.PENDING_TEMPLATE:
            logger.info("检测到旧的 PENDING_TEMPLATE 状态，重置为 PENDING_OUTLINE")
            state_manager.update_task(user_id, conversation_id, corp_id=corp_id, status=TaskStatus.PENDING_OUTLINE)

        # 确认 → 直接生成 PPT（异步执行，不阻塞）
        confirm_keywords = ["确认", "确定", "可以", "好的", "没问题", "同意", "ok", "OK", "yes", "Yes"]
        if any(keyword in text for keyword in confirm_keywords):
            # 如果任务已经在生成中，拒绝重复触发
            if task_state.status == TaskStatus.GENERATING:
                logger.info("PPT正在生成中，忽略重复确认")
                self.reply_text("PPT正在生成中，请稍候...", message)
                return True

            logger.info("用户确认大纲，开始生成PPT（异步）")
            state_manager.update_task(user_id, conversation_id, corp_id=corp_id, status=TaskStatus.GENERATING)

            info = task_state.outline_data
            topic = info.get('topic', '课件')

            # 创建任务ID
            import time as time_module
            task_id = f"ppt_{user_id}_{int(time_module.time() * 1000)}"

            # 获取任务管理器
            from agent.ppt_task_manager import get_ppt_task_manager
            task_manager = get_ppt_task_manager()

            # 定义完成回调
            def on_ppt_complete(task):
                """PPT生成完成回调"""
                try:
                    if task.status.value == "completed" and task.result_path:
                        # 发送PPT文件给用户
                        send_file_message(message, task.result_path, f"{task.result_title}.pptx")

                        # 清理临时文件
                        try:
                            os.remove(task.result_path)
                        except OSError:
                            pass

                        # 更新任务状态
                        state_manager.complete_task(user_id, conversation_id, corp_id)

                        logger.info(f"PPT已发送给用户: {sender_nick}")
                    else:
                        # 生成失败
                        self.reply_text(f"抱歉，PPT生成失败：{task.error_message}", message)
                        state_manager.cancel_task(user_id, conversation_id, corp_id)

                except Exception as e:
                    logger.error(f"发送PPT失败: {e}", exc_info=True)
                    self.reply_text(f"抱歉，发送PPT时出现错误：{str(e)}", message)
                    state_manager.cancel_task(user_id, conversation_id, corp_id)

            # 注册完成回调（必须在submit之前，否则任务可能在回调注册前就完成）
            task_manager.on_complete(task_id, on_ppt_complete)

            # 提交异步任务
            queue_info = task_manager.submit_task(
                task_id=task_id,
                user_id=user_id,
                user_nick=sender_nick,
                conversation_id=conversation_id,
                corp_id=corp_id,
                topic=topic,
                func=generate_ppt_with_master,
                subject=info.get('subject', ''),
                grade=info.get('grade', ''),
                outline_markdown=task_state.outline_markdown,
                page_count=info.get('page_count'),
            )

            # 立即返回，显示队列信息
            queue_position = queue_info.get('queue_position', 0)
            estimated_wait = queue_info.get('estimated_wait_display', '未知')
            running_count = queue_info.get('running_count', 0)
            available_slots = queue_info.get('available_slots', 0)

            # 根据队列情况显示不同提示
            if available_slots > 0 and queue_position == 0:
                # 有空闲线程，立即开始
                status_msg = f"🚀 任务已开始生成！"
                time_msg = f"⏱️ 预计耗时：1-2分钟"
            elif queue_position > 0:
                # 需要排队
                status_msg = f"⏳ 任务已加入队列，当前排队第 {queue_position} 位"
                time_msg = f"⏱️ 预计等待：{estimated_wait}（队列中 {queue_info.get('total_in_queue', 0)} 个任务）"
            else:
                status_msg = f"✅ 任务已提交！"
                time_msg = f"⏱️ 预计耗时：1-2分钟"

            self.reply_text(
                f"{status_msg}\n\n"
                f"📋 主题：{topic}\n"
                f"{time_msg}\n"
                f"📊 当前并发：{running_count}/{queue_info.get('concurrency', {}).get('max', 5)}\n"
                f"📤 生成完成后会自动发送给您\n\n"
                f"💡 您可以继续使用其他功能，不会被阻塞。",
                message,
            )

            return True

        # 修改（异步执行）
        modify_keywords = ["修改", "调整", "更改", "变更", "优化", "改进"]
        if any(keyword in text for keyword in modify_keywords):
            modification = text
            for kw in modify_keywords + ["请", "帮我", "一下"]:
                modification = modification.replace(kw, "")
            modification = modification.strip()

            if not modification:
                self.reply_text("请告诉我您希望如何修改大纲，例如：\n• 修改：增加XXX内容\n• 修改：调整章节顺序", message)
                return True

            self.reply_text("正在根据您的意见修改大纲，请稍候...", message)

            # 异步执行大纲修改
            def modify_outline_task():
                try:
                    outline = self._regenerate_outline(
                        task_state.original_request,
                        task_state.outline_markdown,
                        modification,
                        task_state.outline_data,
                    )

                    state_manager.update_task(
                        user_id, conversation_id,
                        corp_id=corp_id,
                        outline_markdown=outline,
                        status=TaskStatus.PENDING_OUTLINE,
                    )

                    confirm_message = f"""已根据您的意见修改大纲：

{outline}

---
请回复：
• "确认" - 使用此大纲生成PPT
• "修改：[您的修改意见]" - 继续调整
• "取消" - 取消本次生成"""
                    self.reply_text(confirm_message, message)

                except Exception as e:
                    logger.error(f"修改大纲失败: {e}", exc_info=True)
                    self.reply_text(f"抱歉，修改大纲时出现错误：{str(e)}", message)

            from agent.ppt_task_manager import get_ppt_task_manager
            task_manager = get_ppt_task_manager()
            loop = asyncio.get_event_loop()
            loop.run_in_executor(task_manager.executor, modify_outline_task)

            return True

        # 取消
        cancel_keywords = ["取消", "不要了", "算了", "放弃"]
        if any(keyword in text for keyword in cancel_keywords):
            state_manager.cancel_task(user_id, conversation_id, corp_id)
            self.reply_text("好的，已取消本次PPT生成。如有需要，随时告诉我。", message)
            return True

        return False

    def _regenerate_outline(self, original_request, current_outline, feedback, outline_data):
        """根据反馈重新生成大纲"""
        client = __import__('openai').OpenAI(
            api_key=config.OPENAI_API_KEY,
            base_url=config.OPENAI_BASE_URL,
        )

        system_prompt = """你是一个专业的教学内容设计师。请根据用户的反馈，修改PPT大纲。
要求：
1. 保持原有大纲的结构和主要内容
2. 根据用户反馈进行针对性修改
3. 输出完整的修改后大纲（Markdown格式）"""

        response = client.chat.completions.create(
            model=config.OPENAI_MODEL,
            max_tokens=2048,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": f"原始主题：{original_request}\n\n当前大纲：\n{current_outline}\n\n用户修改意见：{feedback}\n\n请生成修改后的大纲。"},
            ],
        )
        return response.choices[0].message.content


def await_search(query: str) -> str:
    """同步等待异步搜索结果"""
    try:
        with concurrent.futures.ThreadPoolExecutor() as executor:
            future = executor.submit(asyncio.run, search_web(query, 5))
            results = future.result(timeout=15)
            if results:
                output = "以下是搜索结果：\n\n"
                for i, r in enumerate(results[:3], 1):
                    output += f"{i}. {r.get('title', '无标题')}\n"
                    if r.get('snippet'):
                        output += f"   摘要：{r['snippet']}\n"
                    output += "\n"
                return output
            return "未找到相关搜索结果。"
    except Exception as e:
        logger.warning(f"搜索失败: {e}")
        return "搜索时出现错误。"


def main():
    logger.info("学校智能助手服务启动（Stream 模式）")

    # 预加载 Embedding 模型（避免首次查询时卡住）
    logger.info("预加载 Embedding 模型...")
    try:
        from agent.knowledge_base_v2 import _get_local_embedding_model
        model = _get_local_embedding_model()
        if model is not None:
            logger.info("Embedding 模型预加载完成")
        else:
            logger.warning("Embedding 模型预加载失败，将使用远程 API")
    except Exception as e:
        logger.warning(f"预加载 Embedding 模型失败: {e}")

    credential = dingtalk_stream.Credential(
        client_id=config.DINGTALK_APP_KEY,
        client_secret=config.DINGTALK_APP_SECRET,
    )
    client = dingtalk_stream.DingTalkStreamClient(credential)

    client.register_callback_handler(
        dingtalk_stream.chatbot.ChatbotMessage.TOPIC,
        SchoolBotHandler(),
    )

    client.start_forever()


if __name__ == "__main__":
    main()
